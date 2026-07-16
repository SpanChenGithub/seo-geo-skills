#!/usr/bin/env python3
"""Validate a content-planner .xlsx workbook and emit a JSON report."""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
import zipfile
from pathlib import Path
from typing import Any, Mapping, Sequence


OPENPYXL_IMPORT_ERROR: ModuleNotFoundError | None = None
try:
    from openpyxl import load_workbook
    from openpyxl.utils.cell import range_boundaries
except ModuleNotFoundError as exc:  # pragma: no cover - depends on caller environment
    OPENPYXL_IMPORT_ERROR = exc

try:
    from build_content_plan_xlsx import (
        ACTION_OPTIONS,
        CONTENT_TYPE_OPTIONS,
        FUNNEL_OPTIONS,
        METHODOLOGY_HEADERS,
        PRIORITY_OPTIONS,
        RESERVED_METHODOLOGY_ROW_COUNT,
        ROADMAP_HEADERS,
        SCHEMA_VERSION,
        SHEET_ORDER,
        STRATEGY_NOTES_HEADERS,
        TOPIC_MAP_HEADERS,
        ArtifactError,
        _keyword_key,
        content_plan_headers,
        load_json_artifact,
        raw_keyword_headers,
    )
except ImportError as exc:  # pragma: no cover - indicates a damaged skill folder
    print(
        json.dumps(
            {
                "status": "error",
                "error": f"cannot import sibling workbook builder: {exc}",
            },
            ensure_ascii=False,
        ),
        file=sys.stderr,
    )
    raise SystemExit(2)


class Validator:
    def __init__(self, workbook_path: Path, source_artifact: Mapping[str, Any] | None = None):
        self.workbook_path = workbook_path
        self.source_artifact = source_artifact
        self.errors: list[str] = []
        self.warnings: list[str] = []
        self.checks: list[str] = []
        self.metadata: dict[str, str] = {}
        self.counts: dict[str, int] = {}
        self.workbook: Any = None

    def error(self, message: str) -> None:
        self.errors.append(message)

    def warning(self, message: str) -> None:
        self.warnings.append(message)

    def checked(self, message: str) -> None:
        self.checks.append(message)

    def validate(self) -> dict[str, Any]:
        if OPENPYXL_IMPORT_ERROR is not None:
            raise RuntimeError(
                "Missing dependency 'openpyxl'. Install it in the active Python "
                "environment (for example: python3 -m pip install openpyxl) and rerun."
            ) from OPENPYXL_IMPORT_ERROR
        if self.workbook_path.suffix.lower() != ".xlsx":
            self.error("workbook must use the .xlsx extension; macro-enabled files are not allowed")
            return self.report()
        if not self.workbook_path.is_file():
            self.error(f"workbook does not exist: {self.workbook_path}")
            return self.report()

        self._check_package_safety()
        try:
            self.workbook = load_workbook(
                self.workbook_path,
                read_only=False,
                data_only=False,
                keep_vba=False,
            )
        except Exception as exc:
            self.error(f"cannot open workbook: {exc}")
            return self.report()

        self._check_sheet_order()
        if "Methodology" in self.workbook.sheetnames:
            self._extract_metadata()
        self._check_independent_market_metadata()
        self._check_headers_layout_and_tables()
        self._check_workbook_properties()
        self._check_formula_safety()
        self._check_content_plan_rules()
        self._check_relationships_and_counts()
        self._compare_source_artifact()
        return self.report()

    def report(self) -> dict[str, Any]:
        valid = not self.errors
        return {
            "status": "valid" if valid else "invalid",
            "valid": valid,
            "schema_version": SCHEMA_VERSION,
            "workbook": str(self.workbook_path.resolve()),
            "metadata": self.metadata,
            "counts": self.counts,
            "checks": self.checks,
            "warnings": self.warnings,
            "errors": self.errors,
        }

    def _check_package_safety(self) -> None:
        try:
            with zipfile.ZipFile(self.workbook_path) as archive:
                members = [name.lower() for name in archive.namelist()]
        except (OSError, zipfile.BadZipFile) as exc:
            self.error(f"workbook is not a valid XLSX package: {exc}")
            return
        forbidden = [
            name
            for name in members
            if "vbaproject" in name or name.endswith(".bin") or "macrosheet" in name
        ]
        if forbidden:
            self.error(f"macros or binary workbook payloads are not allowed: {forbidden}")
        else:
            self.checked("XLSX package contains no macros or binary VBA payloads")

    def _check_sheet_order(self) -> None:
        actual = tuple(self.workbook.sheetnames)
        if actual != SHEET_ORDER:
            self.error(f"sheet order must be {list(SHEET_ORDER)}, found {list(actual)}")
        else:
            self.checked("required six-sheet order is exact")
        if self.workbook.active.title != "Content Plan":
            self.error(
                f"Content Plan must be the active sheet, found {self.workbook.active.title!r}"
            )

    def _extract_metadata(self) -> None:
        ws = self.workbook["Methodology"]
        duplicates: set[str] = set()
        for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
            field, value, _notes = row
            if field is None or str(field).strip() == "":
                continue
            key = str(field).strip()
            if key in self.metadata:
                duplicates.add(key)
                continue
            self.metadata[key] = "" if value is None else _unescape(str(value).strip())
        if duplicates:
            self.error(
                "Methodology contains duplicate reserved/metadata fields: "
                + ", ".join(sorted(duplicates))
            )

    def _check_independent_market_metadata(self) -> None:
        required = (
            "Artifact Schema Version",
            "Project Name",
            "Country Name",
            "Country Code",
            "Volume Label",
            "Language Name",
            "Language Code",
            "Data Source",
            "Content Plan Rows",
            "Raw Keyword Rows",
            "Topic Map Rows",
            "Roadmap Rows",
            "Data Status",
            "Frontier Status",
            "Ahrefs Tools Used",
            "Frontier Items Recorded",
            "Checkpoint Count",
            "Validation Result",
        )
        for field in required:
            if not self.metadata.get(field):
                self.error(f"Methodology metadata field is missing or empty: {field}")
        if self.metadata.get("Artifact Schema Version") not in (None, "", SCHEMA_VERSION):
            self.error(
                f"Artifact Schema Version must be {SCHEMA_VERSION}, "
                f"found {self.metadata.get('Artifact Schema Version')!r}"
            )
        if self.metadata.get("Data Source", "").casefold() != "ahrefs mcp":
            self.error("Data Source must be exactly 'Ahrefs MCP'")
        if self.metadata.get("Ahrefs Tools Used") in (None, "", "N/A"):
            self.error("Ahrefs Tools Used must name at least one MCP tool")
        try:
            frontier_count = int(self.metadata.get("Frontier Items Recorded", "0"))
        except ValueError:
            frontier_count = 0
        if frontier_count < 1:
            self.error("Frontier Items Recorded must be a positive integer")
        for field in ("Country Code", "Language Code"):
            value = self.metadata.get(field, "")
            if any(char in value for char in ",;|/\\\n\r"):
                self.error(f"{field} must identify exactly one market/language, found {value!r}")
        if self.metadata.get("Country Code") and self.metadata.get("Language Code"):
            self.checked("workbook metadata identifies one country and one language")

    def _expected_headers(self) -> dict[str, tuple[str, ...]]:
        pseudo_metadata = {
            "country": {"volume_label": self.metadata.get("Volume Label", "")}
        }
        return {
            "Content Plan": content_plan_headers(pseudo_metadata),
            "Raw Keywords": raw_keyword_headers(pseudo_metadata),
            "Topic Map": TOPIC_MAP_HEADERS,
            "Roadmap": ROADMAP_HEADERS,
            "Strategy Notes": STRATEGY_NOTES_HEADERS,
            "Methodology": METHODOLOGY_HEADERS,
        }

    def _check_headers_layout_and_tables(self) -> None:
        expected_map = self._expected_headers()
        for sheet_name in SHEET_ORDER:
            if sheet_name not in self.workbook.sheetnames:
                continue
            ws = self.workbook[sheet_name]
            expected = expected_map[sheet_name]
            actual = tuple(ws.cell(1, column).value for column in range(1, len(expected) + 1))
            if actual != expected:
                self.error(
                    f"{sheet_name} headers must be {list(expected)}, found {list(actual)}"
                )
            else:
                self.checked(f"{sheet_name} headers match the contract")

            freeze = str(ws.freeze_panes) if ws.freeze_panes is not None else ""
            if freeze != "A2":
                self.error(f"{sheet_name} freeze_panes must be A2, found {freeze!r}")
            if not ws.auto_filter.ref:
                self.error(f"{sheet_name} must have an autofilter")

            tables = list(ws.tables.values())
            if len(tables) != 1:
                self.error(f"{sheet_name} must contain exactly one Excel table")
            else:
                min_col, min_row, max_col, max_row = range_boundaries(tables[0].ref)
                if (min_col, min_row, max_col) != (1, 1, len(expected)):
                    self.error(
                        f"{sheet_name} table must start at A1 and end at header column "
                        f"{len(expected)}, found {tables[0].ref}"
                    )
                effective = _effective_rows(ws, len(expected))
                expected_end = max(2, len(effective) + 1)
                if max_row != expected_end:
                    self.error(
                        f"{sheet_name} table row count does not match populated rows: "
                        f"table ends at {max_row}, expected {expected_end}"
                    )

            for column in range(1, len(expected) + 1):
                cell = ws.cell(1, column)
                color = _rgb(cell.fill.fgColor.rgb)
                if color != "4472C4":
                    self.error(f"{sheet_name} header {cell.coordinate} must use blue fill 4472C4")
                    break
                if not cell.font.bold or _rgb(cell.font.color.rgb if cell.font.color else None) != "FFFFFF":
                    self.error(f"{sheet_name} header row must use bold white text")
                    break

        if "Content Plan" in self.workbook.sheetnames:
            self._check_validations_and_conditional_formatting()

    def _check_validations_and_conditional_formatting(self) -> None:
        ws = self.workbook["Content Plan"]
        expected_columns = {"D": False, "E": False, "N": False, "P": False}
        for validation in ws.data_validations.dataValidation:
            for column in expected_columns:
                if f"{column}2" in validation.cells:
                    expected_columns[column] = True
        for column, found in expected_columns.items():
            if not found:
                self.error(f"Content Plan column {column} is missing list data validation")
        if all(expected_columns.values()):
            self.checked("Content Plan dropdown validation covers Funnel, Content Type, Action, and Priority")

        conditional_ranges = [str(item.sqref) for item in ws.conditional_formatting]
        for column in ("D", "E", "N", "P"):
            if not any(re.search(rf"\b{column}\$?2:{column}", ref) for ref in conditional_ranges):
                self.error(f"Content Plan column {column} is missing conditional color formatting")
        if conditional_ranges:
            self.checked("Content Plan contains color-based conditional formatting")

    def _check_workbook_properties(self) -> None:
        props = self.workbook.properties
        if not props.title or "Content Plan" not in props.title:
            self.error("workbook core title metadata is missing")
        if props.creator != "Span":
            self.error(f"workbook creator must be 'Span', found {props.creator!r}")
        keywords = props.keywords or ""
        expected_tokens = (
            f"schema={SCHEMA_VERSION}",
            f"country={self.metadata.get('Country Code', '')}",
            f"language={self.metadata.get('Language Code', '')}",
            "source=Ahrefs MCP",
        )
        for token in expected_tokens:
            if token not in keywords:
                self.error(f"workbook keywords metadata is missing {token!r}")
        if props.creator == "Span" and props.title:
            self.checked("workbook core metadata is present")

    def _check_formula_safety(self) -> None:
        unsafe: list[str] = []
        for ws in self.workbook.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    value = cell.value
                    if value is None:
                        continue
                    if cell.data_type == "f":
                        unsafe.append(f"{ws.title}!{cell.coordinate} is a formula")
                        continue
                    if isinstance(value, str) and _looks_formula_like(value):
                        unsafe.append(f"{ws.title}!{cell.coordinate} contains unescaped formula-like text")
        if unsafe:
            self.errors.extend(unsafe[:20])
            if len(unsafe) > 20:
                self.error(f"and {len(unsafe) - 20} additional formula-safety violations")
        else:
            self.checked("all populated cells are formula-safe literal values")

    def _check_content_plan_rules(self) -> None:
        if "Content Plan" not in self.workbook.sheetnames:
            return
        ws = self.workbook["Content Plan"]
        rows = _effective_rows(ws, 16)
        self.counts["content_plan"] = len(rows)
        if not rows:
            self.error("Content Plan must contain at least one planned page")
            return
        seen: set[str] = set()
        urls: dict[str, list[tuple[int, str]]] = {}
        for row_number, row in rows:
            primary = _clean_keyword(row[0])
            if not primary:
                self.error(f"Content Plan!A{row_number} Primary Keyword is required")
            elif _keyword_key(primary) in seen:
                self.error(f"Content Plan!A{row_number} duplicates Primary Keyword {primary!r}")
            else:
                seen.add(_keyword_key(primary))
            funnel = _text(row[3]).upper()
            content_type = _text(row[4])
            action = _text(row[13])
            priority = _text(row[15]).upper()
            score = row[14]
            if funnel not in FUNNEL_OPTIONS:
                self.error(f"Content Plan!D{row_number} invalid Funnel {funnel!r}")
            if content_type not in CONTENT_TYPE_OPTIONS:
                self.error(f"Content Plan!E{row_number} invalid Content Type {content_type!r}")
            if action not in ACTION_OPTIONS:
                self.error(f"Content Plan!N{row_number} invalid Action {action!r}")
            if priority not in PRIORITY_OPTIONS:
                self.error(f"Content Plan!P{row_number} invalid Priority {priority!r}")
            if isinstance(score, bool) or not isinstance(score, (int, float)):
                self.error(f"Content Plan!O{row_number} Priority Score must be numeric")
            else:
                expected = "P1" if score >= 75 else "P2" if score >= 50 else "P3"
                if priority != expected:
                    self.error(
                        f"Content Plan row {row_number}: score {score} requires {expected}, found {priority}"
                    )
            for column, label in ((5, "KD"), (6, "Volume"), (7, "Traffic Potential"), (9, "CPC")):
                self._check_metric(row[column], f"Content Plan row {row_number} {label}")
            url = _text(row[12])
            if url:
                urls.setdefault(url.rstrip("/").casefold(), []).append((row_number, action))
                cell = ws.cell(row_number, 13)
                if cell.hyperlink is None:
                    self.error(f"Content Plan!M{row_number} URL must be clickable")
        for url, occurrences in urls.items():
            if len(occurrences) > 1 and not all(action == "Consolidate" for _, action in occurrences):
                self.error(
                    f"Content Plan URL {url!r} is duplicated without all rows marked Consolidate"
                )
        self.checked("Content Plan enums, metric types, unique pages, and priority thresholds checked")

    def _check_metric(self, value: Any, location: str) -> None:
        if value == "N/A":
            return
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            self.error(f"{location} must be numeric or exact text 'N/A', found {value!r}")
            return
        if not math.isfinite(float(value)):
            self.error(f"{location} must be finite")

    def _check_relationships_and_counts(self) -> None:
        needed = set(SHEET_ORDER)
        if not needed.issubset(self.workbook.sheetnames):
            return
        plan_rows = _effective_rows(self.workbook["Content Plan"], 16)
        raw_rows = _effective_rows(self.workbook["Raw Keywords"], 17)
        topic_rows = _effective_rows(self.workbook["Topic Map"], len(TOPIC_MAP_HEADERS))
        roadmap_rows = _effective_rows(self.workbook["Roadmap"], len(ROADMAP_HEADERS))
        strategy_rows = _effective_rows(
            self.workbook["Strategy Notes"], len(STRATEGY_NOTES_HEADERS)
        )
        methodology_rows = _effective_rows(
            self.workbook["Methodology"], len(METHODOLOGY_HEADERS)
        )
        self.counts.update(
            {
                "content_plan": len(plan_rows),
                "raw_keywords": len(raw_rows),
                "topic_map": len(topic_rows),
                "roadmap": len(roadmap_rows),
                "strategy_notes": len(strategy_rows),
                "methodology": len(methodology_rows),
            }
        )

        planned = {_keyword_key(_clean_keyword(row[0])) for _, row in plan_rows if _clean_keyword(row[0])}
        plan_details = {
            _keyword_key(_clean_keyword(row[0])): {
                "topic": _text(row[2]),
                "funnel": _text(row[3]),
                "content_type": _text(row[4]),
                "action": _text(row[13]),
                "priority_score": row[14],
                "priority": _text(row[15]),
            }
            for _, row in plan_rows
            if _clean_keyword(row[0])
        }
        raw_keywords: set[str] = set()
        declared_sources = {
            source.casefold() for source in _split_list(self.metadata.get("Ahrefs Tools Used", ""))
        }
        for row_number, row in raw_rows:
            keyword = _clean_keyword(row[0])
            if not keyword:
                self.error(f"Raw Keywords!A{row_number} Keyword is required")
                continue
            folded = _keyword_key(keyword)
            if folded in raw_keywords:
                self.error(f"Raw Keywords!A{row_number} duplicates keyword {keyword!r}")
            raw_keywords.add(folded)
            country = _text(row[1])
            language = _text(row[2])
            if country != self.metadata.get("Country Code"):
                self.error(
                    f"Raw Keywords!B{row_number} mixes country {country!r} into "
                    f"{self.metadata.get('Country Code')!r} workbook"
                )
            if language != self.metadata.get("Language Code"):
                self.error(
                    f"Raw Keywords!C{row_number} mixes language {language!r} into "
                    f"{self.metadata.get('Language Code')!r} workbook"
                )
            for column, label in ((3, "Volume"), (4, "KD"), (5, "Traffic Potential"), (7, "CPC")):
                self._check_metric(row[column], f"Raw Keywords row {row_number} {label}")
            source_tools = _split_list(row[10])
            if not source_tools:
                self.error(f"Raw Keywords!K{row_number} Source Tool is required")
            for source_tool in source_tools:
                if source_tool.casefold() not in declared_sources:
                    self.error(
                        f"Raw Keywords!K{row_number} contains undeclared Ahrefs tool "
                        f"{source_tool!r}"
                    )
            decision = _text(row[13])
            if decision not in ("Include", "Exclude", "Defer"):
                self.error(f"Raw Keywords!N{row_number} invalid Decision {decision!r}")
            assigned = _clean_keyword(row[15])
            if decision == "Include" and not assigned:
                self.error(
                    f"Raw Keywords!P{row_number} must map included keyword to a planned page"
                )
            if assigned and _keyword_key(assigned) not in planned:
                self.error(
                    f"Raw Keywords!P{row_number} references unplanned page {assigned!r}"
                )
            if _text(row[16]) not in ("Yes", "No"):
                self.error(f"Raw Keywords!Q{row_number} Needs Review must be Yes or No")

        for row_number, row in plan_rows:
            primary = _clean_keyword(row[0])
            if primary and _keyword_key(primary) not in raw_keywords:
                self.error(f"Content Plan!A{row_number} is absent from Raw Keywords: {primary!r}")
            for supporting in _split_list(row[1]):
                if _keyword_key(supporting) not in raw_keywords:
                    self.error(
                        f"Content Plan!B{row_number} supporting keyword is absent from Raw Keywords: "
                        f"{supporting!r}"
                    )

        topic_coverage: set[str] = set()
        for row_number, row in topic_rows:
            primary = _clean_keyword(row[2])
            if _keyword_key(primary) not in planned:
                self.error(f"Topic Map!C{row_number} references unplanned page {primary!r}")
            else:
                topic_coverage.add(_keyword_key(primary))
            url_cell = self.workbook["Topic Map"].cell(row_number, 6)
            if _text(row[5]) and url_cell.hyperlink is None:
                self.error(f"Topic Map!F{row_number} URL must be clickable")
        roadmap_coverage: set[str] = set()
        for row_number, row in roadmap_rows:
            primary = _clean_keyword(row[2])
            if _keyword_key(primary) not in planned:
                self.error(f"Roadmap!C{row_number} references unplanned page {primary!r}")
            else:
                roadmap_coverage.add(_keyword_key(primary))
            content_type = _text(row[5])
            if content_type not in CONTENT_TYPE_OPTIONS:
                self.error(f"Roadmap!F{row_number} invalid Content Type {content_type!r}")
            funnel = _text(row[4])
            action = _text(row[6])
            priority_score = row[7]
            priority = _text(row[8])
            if funnel not in FUNNEL_OPTIONS:
                self.error(f"Roadmap!E{row_number} invalid Funnel {funnel!r}")
            if action not in ACTION_OPTIONS:
                self.error(f"Roadmap!G{row_number} invalid Action {action!r}")
            if priority not in PRIORITY_OPTIONS:
                self.error(f"Roadmap!I{row_number} invalid Priority {priority!r}")
            if isinstance(priority_score, bool) or not isinstance(priority_score, (int, float)):
                self.error(f"Roadmap!H{row_number} Priority Score must be numeric")
            else:
                expected = "P1" if priority_score >= 75 else "P2" if priority_score >= 50 else "P3"
                if priority != expected:
                    self.error(
                        f"Roadmap row {row_number}: score {priority_score} requires {expected}"
                    )
            if _keyword_key(primary) in plan_details:
                roadmap_values = {
                    "topic": _text(row[3]),
                    "funnel": funnel,
                    "content_type": content_type,
                    "action": action,
                    "priority_score": priority_score,
                    "priority": priority,
                }
                for field, expected_value in plan_details[_keyword_key(primary)].items():
                    if roadmap_values[field] != expected_value:
                        self.error(
                            f"Roadmap row {row_number} {field} does not match Content Plan "
                            f"for {primary!r}"
                        )
        missing_topic = planned - topic_coverage
        missing_roadmap = planned - roadmap_coverage
        if missing_topic:
            self.error("planned pages missing from Topic Map: " + ", ".join(sorted(missing_topic)))
        if missing_roadmap:
            self.error("planned pages missing from Roadmap: " + ", ".join(sorted(missing_roadmap)))

        declared = {
            "Content Plan Rows": len(plan_rows),
            "Raw Keyword Rows": len(raw_rows),
            "Topic Map Rows": len(topic_rows),
            "Roadmap Rows": len(roadmap_rows),
        }
        for field, actual in declared.items():
            value = self.metadata.get(field, "")
            try:
                stated = int(value)
            except ValueError:
                self.error(f"Methodology {field} must be an integer, found {value!r}")
                continue
            if stated != actual:
                self.error(f"Methodology {field} says {stated}, workbook contains {actual}")
        self.checked("cross-sheet keyword relationships, page coverage, and declared counts checked")

    def _compare_source_artifact(self) -> None:
        if self.source_artifact is None or self.workbook is None:
            return
        source = self.source_artifact
        meta = source["metadata"]
        expected_metadata = {
            "Country Name": meta["country"]["name"],
            "Country Code": meta["country"]["code"],
            "Volume Label": meta["country"]["volume_label"],
            "Language Name": meta["language"]["name"],
            "Language Code": meta["language"]["code"],
            "Data Source": meta["data_source"],
            "Data Status": meta["data_status"] or "Complete",
            "Frontier Status": meta["frontier_status"] or "N/A",
            "Ahrefs Tools Used": "\n".join(meta["ahrefs_tools_used"]) or "N/A",
            "Frontier Items Recorded": str(len(meta["frontier_items"])),
            "Ahrefs Units Before": str(meta["ahrefs_units_before"]),
            "Ahrefs Units After": str(meta["ahrefs_units_after"]),
            "Site Domain Rating": str(meta["site_domain_rating"]),
        }
        for field, expected in expected_metadata.items():
            if self.metadata.get(field) != expected:
                self.error(
                    f"source JSON {field} is {expected!r}, workbook contains {self.metadata.get(field)!r}"
                )
        source_counts = {
            "content_plan": len(source["content_plan"]),
            "raw_keywords": len(source["raw_keywords"]),
            "topic_map": len(source["topic_map"]),
            "roadmap": len(source["roadmap"]),
            "strategy_notes": len(source["strategy_notes"]),
            "methodology": len(source["methodology"]) + RESERVED_METHODOLOGY_ROW_COUNT,
        }
        for key, expected in source_counts.items():
            if self.counts.get(key) != expected:
                self.error(
                    f"source JSON expects {expected} {key} rows, workbook contains {self.counts.get(key)}"
                )

        plan_rows = _effective_rows(self.workbook["Content Plan"], 16)
        workbook_plan = {_keyword_key(_clean_keyword(row[0])): row for _, row in plan_rows}
        for source_row in source["content_plan"]:
            row = workbook_plan.get(_keyword_key(source_row["primary_keyword"]))
            if row is None:
                continue
            for index, key in ((5, "kd"), (6, "volume"), (7, "traffic_potential"), (9, "cpc_usd")):
                if row[index] != source_row[key]:
                    self.error(
                        f"Content Plan metric changed for {source_row['primary_keyword']!r}: "
                        f"{key} expected {source_row[key]!r}, found {row[index]!r}"
                    )
        raw_rows = _effective_rows(self.workbook["Raw Keywords"], 17)
        workbook_raw = {_keyword_key(_clean_keyword(row[0])): row for _, row in raw_rows}
        for source_row in source["raw_keywords"]:
            row = workbook_raw.get(_keyword_key(source_row["keyword"]))
            if row is None:
                continue
            for index, key in ((3, "volume"), (4, "kd"), (5, "traffic_potential"), (7, "cpc_usd")):
                if row[index] != source_row[key]:
                    self.error(
                        f"Raw Keywords metric changed for {source_row['keyword']!r}: "
                        f"{key} expected {source_row[key]!r}, found {row[index]!r}"
                    )
        self.checked("workbook metadata, counts, N/A values, numeric zeroes, and metrics match source JSON")

    def mark_validation_result(self) -> None:
        if self.workbook is None or "Methodology" not in self.workbook.sheetnames:
            raise ArtifactError("cannot mark validation result without a loaded Methodology sheet")
        worksheet = self.workbook["Methodology"]
        target_row: int | None = None
        for row_number in range(2, worksheet.max_row + 1):
            if _text(worksheet.cell(row_number, 1).value) == "Validation Result":
                target_row = row_number
                break
        if target_row is None:
            raise ArtifactError("Methodology is missing the reserved Validation Result row")
        worksheet.cell(target_row, 2, "Valid")
        worksheet.cell(
            target_row,
            3,
            "Passed validate_content_plan.py against the current workbook",
        )
        self.workbook.save(self.workbook_path)


def _rgb(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    return value[-6:].upper()


def _text(value: Any) -> str:
    return "" if value is None else _unescape(str(value).strip())


def _unescape(value: str) -> str:
    if value.startswith("'") and value[1:].lstrip(" \t\r\n").startswith(("=", "+", "-", "@")):
        return value[1:]
    return value


def _clean_keyword(value: Any) -> str:
    return _text(value)


def _split_list(value: Any) -> list[str]:
    text = _text(value)
    if not text:
        return []
    return [part.strip() for part in re.split(r"[\n;]", text) if part.strip()]


def _looks_formula_like(value: str) -> bool:
    if value.startswith("'"):
        return False
    return value.lstrip(" \t\r\n").startswith(("=", "+", "-", "@"))


def _effective_rows(ws: Any, header_count: int) -> list[tuple[int, tuple[Any, ...]]]:
    result: list[tuple[int, tuple[Any, ...]]] = []
    for row_number in range(2, ws.max_row + 1):
        values = tuple(ws.cell(row_number, column).value for column in range(1, header_count + 1))
        if any(value not in (None, "") for value in values):
            result.append((row_number, values))
    return result


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Validate a content-planner .xlsx and emit a machine-readable JSON report."
    )
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument(
        "--source-json",
        type=Path,
        help="optional source artifact for exact metadata/count/metric comparison",
    )
    parser.add_argument("--output-json", type=Path, help="optional report file")
    parser.add_argument(
        "--mark-valid",
        action="store_true",
        help="after a passing validation, write Valid to Methodology and revalidate",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="explicitly allow replacing an existing JSON report",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    if args.output_json and args.output_json.exists() and not args.overwrite:
        print(
            json.dumps(
                {
                    "status": "error",
                    "error": f"refusing to overwrite existing validation report: {args.output_json}",
                },
                ensure_ascii=False,
            ),
            file=sys.stderr,
        )
        return 2
    try:
        source = load_json_artifact(args.source_json, complete=True) if args.source_json else None
        validator = Validator(args.workbook, source)
        report = validator.validate()
        if args.mark_valid and report["valid"]:
            validator.mark_validation_result()
            validator = Validator(args.workbook, source)
            report = validator.validate()
            report["validation_result_marked"] = bool(
                report["valid"] and report["metadata"].get("Validation Result") == "Valid"
            )
    except (ArtifactError, RuntimeError, OSError) as exc:
        print(json.dumps({"status": "error", "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 2

    rendered = json.dumps(report, ensure_ascii=False, indent=2) + "\n"
    if args.output_json:
        args.output_json.parent.mkdir(parents=True, exist_ok=True)
        args.output_json.write_text(rendered, encoding="utf-8")
    print(rendered, end="")
    return 0 if report["valid"] else 1


if __name__ == "__main__":
    raise SystemExit(main())
