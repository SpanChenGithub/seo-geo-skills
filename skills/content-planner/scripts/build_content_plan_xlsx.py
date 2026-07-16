#!/usr/bin/env python3
"""Build a styled content-planning workbook from a versioned JSON artifact.

This module is deliberately offline. It neither reads environment files nor calls
Ahrefs (or any other network service); it only renders already-collected data.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import sys
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Mapping, Sequence


OPENPYXL_IMPORT_ERROR: ModuleNotFoundError | None = None
try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.workbook.defined_name import DefinedName
except ModuleNotFoundError as exc:  # pragma: no cover - exercised by CLI environments
    OPENPYXL_IMPORT_ERROR = exc


SCHEMA_VERSION = "1.0"
CHECKPOINT_VERSION = "1.0"
CHECKPOINT_BATCH_SIZE = 100
RESERVED_METHODOLOGY_ROW_COUNT = 27
SHEET_ORDER = (
    "Content Plan",
    "Raw Keywords",
    "Topic Map",
    "Roadmap",
    "Strategy Notes",
    "Methodology",
)

CONTENT_PLAN_BASE_HEADERS = (
    "Primary Keyword",
    "Supporting Keywords",
    "Topic",
    "Funnel",
    "Content Type",
    "KD",
    "{volume_header}",
    "Traffic Potential",
    "Search Intent",
    "CPC (USD)",
    "Parent Topic",
    "SERP Features",
    "Existing/Planned URL",
    "Action",
    "Priority Score",
    "Priority",
)

RAW_KEYWORD_HEADERS = (
    "Keyword",
    "Country",
    "Language",
    "Volume",
    "KD",
    "Traffic Potential",
    "Search Intent",
    "CPC (USD)",
    "Parent Topic",
    "SERP Features",
    "Source Tool",
    "Seed/Competitor",
    "SERP Updated",
    "Decision",
    "Decision Reason",
    "Mapped Primary Keyword",
    "Needs Review",
)

TOPIC_MAP_HEADERS = (
    "Topic",
    "Page Level",
    "Primary Keyword",
    "Page Role",
    "Parent Page",
    "Existing/Planned URL",
    "Link Up To",
    "Relevant Cross-Links",
)

ROADMAP_HEADERS = (
    "Phase",
    "Sequence",
    "Primary Keyword",
    "Topic",
    "Funnel",
    "Content Type",
    "Action",
    "Priority Score",
    "Priority",
    "Dependency",
    "Internal Link Targets",
    "Reason",
)

STRATEGY_NOTES_HEADERS = ("Section", "Item", "Details")
METHODOLOGY_HEADERS = ("Field", "Value", "Notes")

FUNNEL_OPTIONS = ("TOFU", "MOFU", "BOFU")
CONTENT_TYPE_OPTIONS = (
    "Tool",
    "Landing Page - Core",
    "Landing Page - Feature",
    "Landing Page - Use Case",
    "Landing Page - Variant",
    "Landing Page - Integration",
    "Landing Page - Comparison",
    "Landing Page - Alternative",
    "Template / Samples",
    "Blog - How to",
    "Blog - Info",
    "Blog - Listicle",
    "Blog - Best",
    "Blog - Review",
    "Blog - Vs",
    "Blog - Examples",
    "Blog - Statistics",
    "Blog - Glossary",
    "Research / Report",
    "Case Study",
)
ACTION_OPTIONS = ("New", "Existing", "Update", "Consolidate")
PRIORITY_OPTIONS = ("P1", "P2", "P3")

HEADER_FILL = PatternFill("solid", fgColor="4472C4") if OPENPYXL_IMPORT_ERROR is None else None
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11) if OPENPYXL_IMPORT_ERROR is None else None
THIN_BLUE = (
    Border(bottom=Side(style="thin", color="2F5597"))
    if OPENPYXL_IMPORT_ERROR is None
    else None
)


class ArtifactError(ValueError):
    """Raised when the JSON artifact violates the public contract."""


SENSITIVE_KEY_PATTERN = re.compile(
    r"(?:^|_)(?:api_?key|password|passwd|authorization|access_?token|"
    r"refresh_?token|private_?key|client_?secret|bearer_?token|secret)(?:$|_)",
    re.IGNORECASE,
)
SENSITIVE_VALUE_PATTERNS = (
    re.compile(r"\bBearer\s+[A-Za-z0-9._~+/=-]{12,}", re.IGNORECASE),
    re.compile(r"\bsk-[A-Za-z0-9_-]{20,}"),
    re.compile(r"\bAIza[0-9A-Za-z_-]{20,}"),
    re.compile(r"-----BEGIN [A-Z ]*PRIVATE KEY-----"),
    re.compile(r"\b[A-Z][A-Z0-9_]*(?:API_KEY|TOKEN|PASSWORD)\s*=\s*\S+"),
)


def require_openpyxl() -> None:
    if OPENPYXL_IMPORT_ERROR is not None:
        raise RuntimeError(
            "Missing dependency 'openpyxl'. Install it in the active Python "
            "environment (for example: python3 -m pip install openpyxl) and rerun."
        ) from OPENPYXL_IMPORT_ERROR


def reject_sensitive_artifact_data(value: Any, path: str = "$") -> None:
    """Reject credential-shaped fields before any artifact value reaches a workbook."""
    if isinstance(value, Mapping):
        for key, item in value.items():
            key_text = str(key)
            normalized_key = re.sub(r"[^A-Za-z0-9]+", "_", key_text).strip("_")
            if SENSITIVE_KEY_PATTERN.search(normalized_key):
                raise ArtifactError(f"credential-like field is forbidden at {path}.{key_text}")
            reject_sensitive_artifact_data(item, f"{path}.{key_text}")
    elif isinstance(value, list):
        for index, item in enumerate(value):
            reject_sensitive_artifact_data(item, f"{path}[{index}]")
    elif isinstance(value, str):
        for pattern in SENSITIVE_VALUE_PATTERNS:
            if pattern.search(value):
                raise ArtifactError(f"credential-like value is forbidden at {path}")


def _required_text(value: Any, path: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ArtifactError(f"{path} must be a non-empty string")
    return value.strip()


def _optional_text(value: Any, path: str) -> str:
    if value is None:
        return ""
    if not isinstance(value, (str, int, float)) or isinstance(value, bool):
        raise ArtifactError(f"{path} must be a string or scalar value")
    return str(value).strip()


def _metric(value: Any, path: str) -> int | float | str:
    """Preserve numeric zero and represent missing Ahrefs metrics as N/A."""
    if value is None or value == "":
        return "N/A"
    if isinstance(value, bool):
        raise ArtifactError(f"{path} must be numeric or 'N/A', not boolean")
    if isinstance(value, (int, float)):
        if not math.isfinite(float(value)):
            raise ArtifactError(f"{path} must be finite")
        return value
    if isinstance(value, str):
        cleaned = value.strip()
        if cleaned.upper() == "N/A":
            return "N/A"
        numeric = cleaned.replace(",", "")
        try:
            parsed = float(numeric)
        except ValueError as exc:
            raise ArtifactError(f"{path} must be numeric or 'N/A'") from exc
        if not math.isfinite(parsed):
            raise ArtifactError(f"{path} must be finite")
        return int(parsed) if parsed.is_integer() else parsed
    raise ArtifactError(f"{path} must be numeric or 'N/A'")


def _score(value: Any, path: str) -> int | float:
    score = _metric(value, path)
    if score == "N/A":
        raise ArtifactError(f"{path} must be a number from 0 to 100")
    numeric = float(score)
    if numeric < 0 or numeric > 100:
        raise ArtifactError(f"{path} must be between 0 and 100")
    return int(numeric) if numeric.is_integer() else numeric


def _text_list(value: Any, path: str) -> list[str]:
    if value is None or value == "":
        return []
    if isinstance(value, str):
        return [part.strip() for part in re.split(r"[\n;]", value) if part.strip()]
    if not isinstance(value, list):
        raise ArtifactError(f"{path} must be an array of strings or a string")
    result: list[str] = []
    for index, item in enumerate(value):
        if not isinstance(item, (str, int, float)) or isinstance(item, bool):
            raise ArtifactError(f"{path}[{index}] must be a scalar string value")
        text = str(item).strip()
        if text:
            result.append(text)
    return result


def _keyword_key(value: str) -> str:
    """Build the documented comparison key without changing display spelling."""
    normalized = unicodedata.normalize("NFKC", value)
    return " ".join(normalized.split()).casefold()


def _audit_text(value: Any, path: str) -> str:
    if isinstance(value, (Mapping, list)):
        try:
            return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        except (TypeError, ValueError) as exc:
            raise ArtifactError(f"{path} must contain JSON-serializable audit data") from exc
    return _optional_text(value, path)


def _normalize_frontier_items(value: Any, tools: Sequence[str]) -> list[dict[str, Any]]:
    if not isinstance(value, list) or not value:
        raise ArtifactError("metadata.frontier_items must be a non-empty array")
    declared_tools = {tool.casefold() for tool in tools}
    result: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    allowed_statuses = {"Queued", "Completed", "Exhausted", "Failed", "Partial"}
    for index, item in enumerate(value):
        path = f"metadata.frontier_items[{index}]"
        if not isinstance(item, Mapping):
            raise ArtifactError(f"{path} must be an object")
        frontier_id = _required_text(item.get("frontier_id"), f"{path}.frontier_id")
        if frontier_id in seen_ids:
            raise ArtifactError(f"{path}.frontier_id duplicates another frontier item")
        seen_ids.add(frontier_id)
        source_tool = _required_text(item.get("source_tool"), f"{path}.source_tool")
        if source_tool.casefold() not in declared_tools:
            raise ArtifactError(
                f"{path}.source_tool is not listed in metadata.ahrefs_tools_used"
            )
        status = _required_text(item.get("status"), f"{path}.status").title()
        if status not in allowed_statuses:
            raise ArtifactError(
                f"{path}.status must be one of {', '.join(sorted(allowed_statuses))}"
            )
        result.append(
            {
                "frontier_id": frontier_id,
                "source_tool": source_tool,
                "target": _required_text(item.get("target"), f"{path}.target"),
                "mode": _optional_text(item.get("mode"), f"{path}.mode"),
                "filters": _audit_text(item.get("filters"), f"{path}.filters"),
                "selected_fields": _text_list(
                    item.get("selected_fields"), f"{path}.selected_fields"
                ),
                "status": status,
                "returned_rows": _metric(item.get("returned_rows"), f"{path}.returned_rows"),
                "new_unique_rows": _metric(
                    item.get("new_unique_rows"), f"{path}.new_unique_rows"
                ),
                "duplicate_rows": _metric(
                    item.get("duplicate_rows"), f"{path}.duplicate_rows"
                ),
                "included_rows": _metric(
                    item.get("included_rows"), f"{path}.included_rows"
                ),
                "excluded_rows": _metric(
                    item.get("excluded_rows"), f"{path}.excluded_rows"
                ),
                "units_before": _metric(item.get("units_before"), f"{path}.units_before"),
                "units_after": _metric(item.get("units_after"), f"{path}.units_after"),
                "collected_at": _optional_text(item.get("collected_at"), f"{path}.collected_at"),
            }
        )
    return result


def _priority_for_score(score: float) -> str:
    if score >= 75:
        return "P1"
    if score >= 50:
        return "P2"
    return "P3"


def _normalize_metadata(raw: Any) -> dict[str, Any]:
    if not isinstance(raw, Mapping):
        raise ArtifactError("metadata must be an object")

    project_name = _required_text(raw.get("project_name"), "metadata.project_name")
    country = raw.get("country")
    language = raw.get("language")
    if not isinstance(country, Mapping):
        raise ArtifactError("metadata.country must be an object")
    if not isinstance(language, Mapping):
        raise ArtifactError("metadata.language must be an object")

    country_name = _required_text(country.get("name"), "metadata.country.name")
    country_code = _required_text(country.get("code"), "metadata.country.code").upper()
    if any(char in country_code for char in ",;|/\\\n\r"):
        raise ArtifactError("metadata.country.code must identify exactly one country")
    language_name = _required_text(language.get("name"), "metadata.language.name")
    language_code = _required_text(language.get("code"), "metadata.language.code")
    if any(char in language_code for char in ",;|/\\\n\r"):
        raise ArtifactError("metadata.language.code must identify exactly one language")
    volume_label = _optional_text(
        country.get("volume_label") or country_code or country_name,
        "metadata.country.volume_label",
    )
    data_source = _required_text(raw.get("data_source"), "metadata.data_source")
    if data_source.casefold() != "ahrefs mcp":
        raise ArtifactError("metadata.data_source must be exactly 'Ahrefs MCP'")

    generated_at = _optional_text(raw.get("generated_at"), "metadata.generated_at")
    if not generated_at:
        generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()

    ahrefs_tools_used = _text_list(
        raw.get("ahrefs_tools_used"), "metadata.ahrefs_tools_used"
    )
    if not ahrefs_tools_used:
        raise ArtifactError("metadata.ahrefs_tools_used must contain at least one MCP tool")
    frontier_items = _normalize_frontier_items(raw.get("frontier_items"), ahrefs_tools_used)

    return {
        "project_name": project_name,
        "project_type": _optional_text(raw.get("project_type"), "metadata.project_type"),
        "site_url": _optional_text(raw.get("site_url"), "metadata.site_url"),
        "country": {
            "name": country_name,
            "code": country_code,
            "volume_label": volume_label,
        },
        "language": {"name": language_name, "code": language_code},
        "data_source": "Ahrefs MCP",
        "generated_at": generated_at,
        "data_date": _optional_text(raw.get("data_date"), "metadata.data_date"),
        "run_id": _optional_text(raw.get("run_id"), "metadata.run_id"),
        "checkpoint_count": _optional_text(
            raw.get("checkpoint_count"), "metadata.checkpoint_count"
        ),
        "validation_result": _optional_text(
            raw.get("validation_result"), "metadata.validation_result"
        ),
        "data_status": _optional_text(raw.get("data_status"), "metadata.data_status"),
        "frontier_status": _optional_text(
            raw.get("frontier_status"), "metadata.frontier_status"
        ),
        "ahrefs_tools_used": ahrefs_tools_used,
        "frontier_items": frontier_items,
        "ahrefs_units_before": _metric(
            raw.get("ahrefs_units_before"), "metadata.ahrefs_units_before"
        ),
        "ahrefs_units_after": _metric(
            raw.get("ahrefs_units_after"), "metadata.ahrefs_units_after"
        ),
        "site_domain_rating": _metric(
            raw.get("site_domain_rating"), "metadata.site_domain_rating"
        ),
    }


def _require_rows(data: Mapping[str, Any], key: str, *, allow_empty: bool = False) -> list[Mapping[str, Any]]:
    value = data.get(key)
    if not isinstance(value, list):
        raise ArtifactError(f"{key} must be an array")
    if not allow_empty and not value:
        raise ArtifactError(f"{key} must contain at least one row")
    for index, row in enumerate(value):
        if not isinstance(row, Mapping):
            raise ArtifactError(f"{key}[{index}] must be an object")
    return value


def _normalize_content_plan(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for index, row in enumerate(rows):
        path = f"content_plan[{index}]"
        primary = _required_text(row.get("primary_keyword"), f"{path}.primary_keyword")
        folded = _keyword_key(primary)
        if folded in seen:
            raise ArtifactError(f"{path}.primary_keyword duplicates another planned page: {primary!r}")
        seen.add(folded)
        funnel = _required_text(row.get("funnel"), f"{path}.funnel").upper()
        content_type = _required_text(row.get("content_type"), f"{path}.content_type")
        action = _required_text(row.get("action"), f"{path}.action")
        priority = _required_text(row.get("priority"), f"{path}.priority").upper()
        score = _score(row.get("priority_score"), f"{path}.priority_score")
        if funnel not in FUNNEL_OPTIONS:
            raise ArtifactError(f"{path}.funnel must be one of {', '.join(FUNNEL_OPTIONS)}")
        if content_type not in CONTENT_TYPE_OPTIONS:
            raise ArtifactError(f"{path}.content_type is not an allowed content type")
        if action not in ACTION_OPTIONS:
            raise ArtifactError(f"{path}.action must be one of {', '.join(ACTION_OPTIONS)}")
        if priority not in PRIORITY_OPTIONS:
            raise ArtifactError(f"{path}.priority must be one of {', '.join(PRIORITY_OPTIONS)}")
        expected = _priority_for_score(float(score))
        if priority != expected:
            raise ArtifactError(
                f"{path}.priority must be {expected} for priority_score {score}, not {priority}"
            )
        provisional = row.get("score_is_provisional", False)
        if not isinstance(provisional, bool):
            raise ArtifactError(f"{path}.score_is_provisional must be a boolean")
        score_breakdown = row.get("score_breakdown", {})
        if score_breakdown is None:
            score_breakdown = {}
        if not isinstance(score_breakdown, Mapping):
            raise ArtifactError(f"{path}.score_breakdown must be an object")
        normalized.append(
            {
                "primary_keyword": primary,
                "supporting_keywords": _text_list(
                    row.get("supporting_keywords"), f"{path}.supporting_keywords"
                ),
                "topic": _required_text(row.get("topic"), f"{path}.topic"),
                "funnel": funnel,
                "content_type": content_type,
                "kd": _metric(row.get("kd"), f"{path}.kd"),
                "volume": _metric(row.get("volume"), f"{path}.volume"),
                "traffic_potential": _metric(
                    row.get("traffic_potential"), f"{path}.traffic_potential"
                ),
                "search_intent": _text_list(row.get("search_intent"), f"{path}.search_intent"),
                "cpc_usd": _metric(row.get("cpc_usd"), f"{path}.cpc_usd"),
                "parent_topic": _optional_text(row.get("parent_topic"), f"{path}.parent_topic"),
                "serp_features": _text_list(row.get("serp_features"), f"{path}.serp_features"),
                "url": _optional_text(row.get("url"), f"{path}.url"),
                "action": action,
                "priority_score": score,
                "priority": priority,
                "score_is_provisional": provisional,
                "score_breakdown": dict(score_breakdown),
            }
        )
    return sorted(
        normalized,
        key=lambda row: (
            PRIORITY_OPTIONS.index(row["priority"]),
            -float(row["priority_score"]),
            row["topic"].casefold(),
            -(float(row["volume"]) if isinstance(row["volume"], (int, float)) else -1.0),
            row["primary_keyword"].casefold(),
        ),
    )


def _normalize_raw_keywords(
    rows: Sequence[Mapping[str, Any]], metadata: Mapping[str, Any]
) -> list[dict[str, Any]]:
    normalized: list[dict[str, Any]] = []
    seen: set[str] = set()
    for index, row in enumerate(rows):
        path = f"raw_keywords[{index}]"
        keyword = _required_text(row.get("keyword"), f"{path}.keyword")
        folded = _keyword_key(keyword)
        if folded in seen:
            raise ArtifactError(f"{path}.keyword duplicates another raw keyword: {keyword!r}")
        seen.add(folded)
        country_input = _optional_text(row.get("country"), f"{path}.country")
        language_input = _optional_text(row.get("language"), f"{path}.language")
        accepted_countries = {
            metadata["country"]["code"].casefold(),
            metadata["country"]["name"].casefold(),
        }
        accepted_languages = {
            metadata["language"]["code"].casefold(),
            metadata["language"]["name"].casefold(),
        }
        if country_input and country_input.casefold() not in accepted_countries:
            raise ArtifactError(
                f"{path}.country must match the workbook country "
                f"{metadata['country']['code']!r}"
            )
        if language_input and language_input.casefold() not in accepted_languages:
            raise ArtifactError(
                f"{path}.language must match the workbook language "
                f"{metadata['language']['code']!r}"
            )
        sources = _text_list(
            row.get("source_tool", row.get("ahrefs_source")), f"{path}.source_tool"
        )
        if not sources:
            raise ArtifactError(f"{path}.source_tool must name at least one Ahrefs MCP tool")
        declared_sources = {tool.casefold() for tool in metadata["ahrefs_tools_used"]}
        undeclared_sources = [
            source for source in sources if source.casefold() not in declared_sources
        ]
        if undeclared_sources:
            raise ArtifactError(
                f"{path}.source_tool contains tools not listed in metadata.ahrefs_tools_used: "
                + ", ".join(undeclared_sources)
            )
        evidenced_sources = {
            item["source_tool"].casefold()
            for item in metadata["frontier_items"]
            if item["status"] in {"Completed", "Partial", "Exhausted"}
        }
        unevidenced_sources = [
            source for source in sources if source.casefold() not in evidenced_sources
        ]
        if unevidenced_sources:
            raise ArtifactError(
                f"{path}.source_tool has no completed, partial, or exhausted frontier item: "
                + ", ".join(unevidenced_sources)
            )
        needs_review_raw = row.get("needs_review", False)
        if isinstance(needs_review_raw, bool):
            needs_review = "Yes" if needs_review_raw else "No"
        elif isinstance(needs_review_raw, str) and needs_review_raw.strip().casefold() in {
            "yes",
            "no",
        }:
            needs_review = needs_review_raw.strip().title()
        else:
            raise ArtifactError(f"{path}.needs_review must be a boolean or 'Yes'/'No'")
        decision = _required_text(row.get("decision"), f"{path}.decision").title()
        if decision not in ("Include", "Exclude", "Defer"):
            raise ArtifactError(f"{path}.decision must be Include, Exclude, or Defer")
        mapped = _optional_text(
            row.get("mapped_primary_keyword", row.get("assigned_primary_keyword")),
            f"{path}.mapped_primary_keyword",
        )
        if decision == "Include" and not mapped:
            raise ArtifactError(
                f"{path}.mapped_primary_keyword is required when decision is Include"
            )
        normalized.append(
            {
                "keyword": keyword,
                "country": metadata["country"]["code"],
                "language": metadata["language"]["code"],
                "volume": _metric(row.get("volume"), f"{path}.volume"),
                "kd": _metric(row.get("kd"), f"{path}.kd"),
                "traffic_potential": _metric(
                    row.get("traffic_potential"), f"{path}.traffic_potential"
                ),
                "search_intent": _text_list(row.get("search_intent"), f"{path}.search_intent"),
                "cpc_usd": _metric(row.get("cpc_usd"), f"{path}.cpc_usd"),
                "parent_topic": _optional_text(row.get("parent_topic"), f"{path}.parent_topic"),
                "serp_features": _text_list(row.get("serp_features"), f"{path}.serp_features"),
                "source_tool": sources,
                "seed_or_competitor": _optional_text(
                    row.get("seed_or_competitor", row.get("seed_keyword")),
                    f"{path}.seed_or_competitor",
                ),
                "serp_updated": _optional_text(
                    row.get("serp_updated", row.get("serp_updated_at")),
                    f"{path}.serp_updated",
                ),
                "decision": decision,
                "decision_reason": _optional_text(
                    row.get("decision_reason"), f"{path}.decision_reason"
                ),
                "mapped_primary_keyword": mapped,
                "needs_review": needs_review,
            }
        )
    return normalized


def _normalize_topic_map(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        path = f"topic_map[{index}]"
        result.append(
            {
                "topic": _required_text(row.get("topic"), f"{path}.topic"),
                "page_level": _required_text(row.get("page_level"), f"{path}.page_level"),
                "primary_keyword": _required_text(
                    row.get("primary_keyword"), f"{path}.primary_keyword"
                ),
                "page_role": _required_text(
                    row.get("page_role", row.get("role")), f"{path}.page_role"
                ),
                "parent_page": _optional_text(
                    row.get("parent_page", row.get("parent_page_keyword")),
                    f"{path}.parent_page",
                ),
                "url": _optional_text(row.get("url"), f"{path}.url"),
                "link_up_to": _text_list(
                    row.get("link_up_to", row.get("internal_link_to")), f"{path}.link_up_to"
                ),
                "relevant_cross_links": _text_list(
                    row.get("relevant_cross_links"), f"{path}.relevant_cross_links"
                ),
            }
        )
    return result


def _normalize_roadmap(rows: Sequence[Mapping[str, Any]]) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    for index, row in enumerate(rows):
        path = f"roadmap[{index}]"
        sequence = row.get("sequence", row.get("order", index + 1))
        if isinstance(sequence, bool) or not isinstance(sequence, (int, float, str)):
            raise ArtifactError(f"{path}.sequence must be a scalar")
        funnel = _required_text(row.get("funnel"), f"{path}.funnel").upper()
        content_type = _required_text(row.get("content_type"), f"{path}.content_type")
        action = _required_text(row.get("action"), f"{path}.action")
        priority_score = _score(row.get("priority_score"), f"{path}.priority_score")
        priority = _required_text(row.get("priority"), f"{path}.priority").upper()
        if funnel not in FUNNEL_OPTIONS:
            raise ArtifactError(f"{path}.funnel must be one of {', '.join(FUNNEL_OPTIONS)}")
        if content_type not in CONTENT_TYPE_OPTIONS:
            raise ArtifactError(f"{path}.content_type is not an allowed content type")
        if action not in ACTION_OPTIONS:
            raise ArtifactError(f"{path}.action must be one of {', '.join(ACTION_OPTIONS)}")
        expected_priority = _priority_for_score(float(priority_score))
        if priority != expected_priority:
            raise ArtifactError(
                f"{path}.priority must be {expected_priority} for score {priority_score}"
            )
        result.append(
            {
                "phase": _required_text(row.get("phase"), f"{path}.phase"),
                "sequence": sequence,
                "primary_keyword": _required_text(
                    row.get("primary_keyword"), f"{path}.primary_keyword"
                ),
                "topic": _required_text(row.get("topic"), f"{path}.topic"),
                "funnel": funnel,
                "content_type": content_type,
                "action": action,
                "priority_score": priority_score,
                "priority": priority,
                "dependency": _text_list(row.get("dependency"), f"{path}.dependency"),
                "internal_link_targets": _text_list(
                    row.get("internal_link_targets", row.get("internal_link_to")),
                    f"{path}.internal_link_targets",
                ),
                "reason": _required_text(
                    row.get("reason", row.get("rationale")), f"{path}.reason"
                ),
            }
        )
    return result


def _normalize_three_column_rows(
    rows: Sequence[Mapping[str, Any]],
    section_key: str,
    keys: tuple[str, str, str],
) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    for index, row in enumerate(rows):
        path = f"{section_key}[{index}]"
        result.append(
            {
                keys[0]: _required_text(row.get(keys[0]), f"{path}.{keys[0]}"),
                keys[1]: _required_text(row.get(keys[1]), f"{path}.{keys[1]}"),
                keys[2]: _required_text(row.get(keys[2]), f"{path}.{keys[2]}"),
            }
        )
    return result


def normalize_artifact(data: Any, *, complete: bool = True) -> dict[str, Any]:
    if not isinstance(data, Mapping):
        raise ArtifactError("JSON root must be an object")
    reject_sensitive_artifact_data(data)
    if str(data.get("schema_version", "")) != SCHEMA_VERSION:
        raise ArtifactError(f"schema_version must be {SCHEMA_VERSION!r}")
    metadata = _normalize_metadata(data.get("metadata"))
    raw_keywords = _normalize_raw_keywords(
        _require_rows(data, "raw_keywords", allow_empty=not complete), metadata
    )
    if not complete:
        return {
            "schema_version": SCHEMA_VERSION,
            "metadata": metadata,
            "raw_keywords": raw_keywords,
        }

    content_plan = _normalize_content_plan(_require_rows(data, "content_plan"))
    topic_map = _normalize_topic_map(_require_rows(data, "topic_map"))
    roadmap = _normalize_roadmap(_require_rows(data, "roadmap"))
    strategy_notes = _normalize_three_column_rows(
        _require_rows(data, "strategy_notes"),
        "strategy_notes",
        ("section", "item", "details"),
    )
    methodology = _normalize_three_column_rows(
        _require_rows(data, "methodology"),
        "methodology",
        ("field", "value", "notes"),
    )

    planned = {_keyword_key(row["primary_keyword"]): row for row in content_plan}
    raw = {_keyword_key(row["keyword"]): row["keyword"] for row in raw_keywords}
    for row in content_plan:
        related = [row["primary_keyword"], *row["supporting_keywords"]]
        for keyword in related:
            if _keyword_key(keyword) not in raw:
                raise ArtifactError(
                    f"planned keyword {keyword!r} is missing from raw_keywords"
                )
    for index, row in enumerate(raw_keywords):
        assigned = row["mapped_primary_keyword"]
        if assigned and _keyword_key(assigned) not in planned:
            raise ArtifactError(
                f"raw_keywords[{index}].mapped_primary_keyword does not match a planned page"
            )
    for index, row in enumerate(topic_map):
        if _keyword_key(row["primary_keyword"]) not in planned:
            raise ArtifactError(f"topic_map[{index}].primary_keyword is not in content_plan")
    for index, row in enumerate(roadmap):
        key = _keyword_key(row["primary_keyword"])
        if key not in planned:
            raise ArtifactError(f"roadmap[{index}].primary_keyword is not in content_plan")
        plan_row = planned[key]
        for field in ("topic", "funnel", "content_type", "action", "priority_score", "priority"):
            if row[field] != plan_row[field]:
                raise ArtifactError(
                    f"roadmap[{index}].{field} must match the Content Plan row for "
                    f"{row['primary_keyword']!r}"
                )
    topic_coverage = {_keyword_key(row["primary_keyword"]) for row in topic_map}
    roadmap_coverage = {_keyword_key(row["primary_keyword"]) for row in roadmap}
    missing_topic = set(planned) - topic_coverage
    missing_roadmap = set(planned) - roadmap_coverage
    if missing_topic:
        raise ArtifactError("topic_map is missing planned pages: " + ", ".join(sorted(missing_topic)))
    if missing_roadmap:
        raise ArtifactError("roadmap is missing planned pages: " + ", ".join(sorted(missing_roadmap)))

    return {
        "schema_version": SCHEMA_VERSION,
        "metadata": metadata,
        "content_plan": content_plan,
        "raw_keywords": raw_keywords,
        "topic_map": topic_map,
        "roadmap": roadmap,
        "strategy_notes": strategy_notes,
        "methodology": methodology,
    }


def load_json_artifact(path: Path, *, complete: bool = True) -> dict[str, Any]:
    try:
        text = path.read_text(encoding="utf-8-sig")
    except OSError as exc:
        raise ArtifactError(f"cannot read JSON artifact {path}: {exc}") from exc
    try:
        data = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ArtifactError(f"invalid UTF-8 JSON in {path}: {exc}") from exc
    return normalize_artifact(data, complete=complete)


def escape_formula_injection(value: Any) -> Any:
    """Force formula-like text to remain literal text in spreadsheet cells."""
    if not isinstance(value, str):
        return value
    stripped = value.lstrip(" \t\r\n")
    if stripped.startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


def _join_lines(value: Any) -> Any:
    if isinstance(value, list):
        return "\n".join(value)
    return value


def _volume_header(metadata: Mapping[str, Any]) -> str:
    return f"Volume ({metadata['country']['volume_label']})"


def content_plan_headers(metadata: Mapping[str, Any]) -> tuple[str, ...]:
    volume = _volume_header(metadata)
    return tuple(header.format(volume_header=volume) for header in CONTENT_PLAN_BASE_HEADERS)


def raw_keyword_headers(metadata: Mapping[str, Any]) -> tuple[str, ...]:
    del metadata
    return RAW_KEYWORD_HEADERS


def _content_plan_values(row: Mapping[str, Any]) -> list[Any]:
    return [
        row["primary_keyword"],
        _join_lines(row["supporting_keywords"]),
        row["topic"],
        row["funnel"],
        row["content_type"],
        row["kd"],
        row["volume"],
        row["traffic_potential"],
        _join_lines(row["search_intent"]),
        row["cpc_usd"],
        row["parent_topic"],
        _join_lines(row["serp_features"]),
        row["url"],
        row["action"],
        row["priority_score"],
        row["priority"],
    ]


def _raw_keyword_values(row: Mapping[str, Any]) -> list[Any]:
    return [
        row["keyword"],
        row["country"],
        row["language"],
        row["volume"],
        row["kd"],
        row["traffic_potential"],
        _join_lines(row["search_intent"]),
        row["cpc_usd"],
        row["parent_topic"],
        _join_lines(row["serp_features"]),
        _join_lines(row["source_tool"]),
        row["seed_or_competitor"],
        row["serp_updated"],
        row["decision"],
        row["decision_reason"],
        row["mapped_primary_keyword"],
        row["needs_review"],
    ]


def _topic_map_values(row: Mapping[str, Any]) -> list[Any]:
    return [
        row["topic"],
        row["page_level"],
        row["primary_keyword"],
        row["page_role"],
        row["parent_page"],
        row["url"],
        _join_lines(row["link_up_to"]),
        _join_lines(row["relevant_cross_links"]),
    ]


def _roadmap_values(row: Mapping[str, Any]) -> list[Any]:
    return [
        row["phase"],
        row["sequence"],
        row["primary_keyword"],
        row["topic"],
        row["funnel"],
        row["content_type"],
        row["action"],
        row["priority_score"],
        row["priority"],
        _join_lines(row["dependency"]),
        _join_lines(row["internal_link_targets"]),
        row["reason"],
    ]


def _strategy_values(row: Mapping[str, Any]) -> list[Any]:
    return [row["section"], row["item"], row["details"]]


def _methodology_values(row: Mapping[str, Any]) -> list[Any]:
    return [row["field"], row["value"], row["notes"]]


def _methodology_rows(artifact: Mapping[str, Any]) -> list[dict[str, str]]:
    meta = artifact["metadata"]
    reserved = [
        ("Artifact Schema Version", SCHEMA_VERSION, "Workbook contract version"),
        ("Project Name", meta["project_name"], "One country and language per workbook"),
        ("Project Type", meta["project_type"] or "N/A", "New or existing website"),
        ("Website URL", meta["site_url"] or "N/A", "No live-site changes are made"),
        ("Country Name", meta["country"]["name"], "Required analysis market"),
        ("Country Code", meta["country"]["code"], "Exactly one country"),
        ("Volume Label", meta["country"]["volume_label"], "Used in Volume column headers"),
        ("Language Name", meta["language"]["name"], "Required content language"),
        ("Language Code", meta["language"]["code"], "Exactly one language"),
        ("Data Source", meta["data_source"], "Keyword facts and metrics only from Ahrefs MCP"),
        ("Data Status", meta["data_status"] or "Complete", "Mark Partial when research stopped early"),
        (
            "Frontier Status",
            meta["frontier_status"] or "N/A",
            "Why discovery stopped or whether the frontier was exhausted",
        ),
        (
            "Ahrefs Tools Used",
            "\n".join(meta["ahrefs_tools_used"]) or "N/A",
            "MCP methods used for this country run",
        ),
        (
            "Frontier Items Recorded",
            str(len(meta["frontier_items"])),
            "Auditable non-replayed Ahrefs request partitions",
        ),
        (
            "Ahrefs Units Before",
            str(meta["ahrefs_units_before"]),
            "Optional units snapshot; never store credentials or account IDs",
        ),
        (
            "Ahrefs Units After",
            str(meta["ahrefs_units_after"]),
            "Optional units snapshot; never store credentials or account IDs",
        ),
        (
            "Site Domain Rating",
            str(meta["site_domain_rating"]),
            "Ahrefs DR for existing sites when available",
        ),
        ("Generated At", meta["generated_at"], "ISO-8601 timestamp"),
        ("Data Date", meta["data_date"] or "N/A", "Ahrefs data date when available"),
        ("Run ID", meta["run_id"] or "N/A", "Optional resumable-run identifier"),
        ("Content Plan Rows", str(len(artifact["content_plan"])), "Planned pages"),
        ("Raw Keyword Rows", str(len(artifact["raw_keywords"])), "Deduplicated Ahrefs candidates"),
        ("Topic Map Rows", str(len(artifact["topic_map"])), "Topic architecture rows"),
        ("Roadmap Rows", str(len(artifact["roadmap"])), "Publishing roadmap rows"),
        (
            "Checkpoint Count",
            meta["checkpoint_count"] or str(math.ceil(len(artifact["raw_keywords"]) / CHECKPOINT_BATCH_SIZE)),
            "One non-overwriting checkpoint per 100 raw rows",
        ),
        (
            "Validation Result",
            meta["validation_result"] or "Pending",
            "Use validate_content_plan.py --mark-valid for the final workbook",
        ),
        ("Author", "Span", "Workbook generator metadata"),
    ]
    rows = [{"field": field, "value": value, "notes": notes} for field, value, notes in reserved]
    if len(rows) != RESERVED_METHODOLOGY_ROW_COUNT:
        raise ArtifactError("internal methodology metadata row count is inconsistent")
    reserved_fields = {row["field"].casefold() for row in rows}
    for row in artifact["methodology"]:
        if row["field"].casefold() in reserved_fields:
            raise ArtifactError(
                f"methodology field {row['field']!r} is reserved for workbook metadata"
            )
    rows.extend(artifact["methodology"])
    return rows


def _add_table(ws: Any, table_name: str, header_count: int, data_count: int) -> None:
    end_col = get_column_letter(header_count)
    end_row = max(2, data_count + 1)
    table = Table(displayName=table_name, ref=f"A1:{end_col}{end_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.auto_filter.ref = f"A1:{end_col}{end_row}"


def _style_sheet(ws: Any, headers: Sequence[str], data_count: int, table_name: str) -> None:
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 26
    ws.print_title_rows = "1:1"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(2, data_count + 1)}"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.outlinePr.summaryBelow = True

    for cell in ws[1][: len(headers)]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BLUE
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2, max_row=max(2, data_count + 1), max_col=len(headers)):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row[0].row].height = 30

    widths = {
        "Primary Keyword": 30,
        "Supporting Keywords": 34,
        "Keyword": 30,
        "Topic": 22,
        "Funnel": 12,
        "Content Type": 26,
        "Search Intent": 24,
        "Parent Topic": 24,
        "SERP Features": 28,
        "Existing/Planned URL": 42,
        "Action": 14,
        "Priority Score": 15,
        "Priority": 12,
        "Source Tool": 22,
        "Seed/Competitor": 28,
        "Decision Reason": 38,
        "Mapped Primary Keyword": 30,
        "SERP Updated": 22,
        "Relevant Cross-Links": 36,
        "Internal Link Targets": 36,
        "Reason": 45,
        "Details": 80,
        "Value": 42,
    }
    for index, header in enumerate(headers, start=1):
        default = 16 if not header.startswith("Volume (") else 16
        ws.column_dimensions[get_column_letter(index)].width = widths.get(header, default)

    _add_table(ws, table_name, len(headers), data_count)


def _write_sheet(
    ws: Any,
    headers: Sequence[str],
    rows: Sequence[Mapping[str, Any]],
    value_function: Any,
    table_name: str,
) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append([escape_formula_injection(value) for value in value_function(row)])
    if not rows:
        ws.append([None] * len(headers))
    _style_sheet(ws, headers, len(rows), table_name)


def _add_defined_name(workbook: Any, name: str, reference: str) -> None:
    defined_name = DefinedName(name, attr_text=reference)
    if hasattr(workbook.defined_names, "add"):
        workbook.defined_names.add(defined_name)
    else:  # openpyxl 3.0 compatibility
        workbook.defined_names.append(defined_name)


def _add_validation_options(workbook: Any) -> None:
    ws = workbook["Methodology"]
    option_groups = (
        (5, "Funnel Options", FUNNEL_OPTIONS, "_CP_FunnelOptions"),
        (6, "Content Type Options", CONTENT_TYPE_OPTIONS, "_CP_ContentTypeOptions"),
        (7, "Action Options", ACTION_OPTIONS, "_CP_ActionOptions"),
        (8, "Priority Options", PRIORITY_OPTIONS, "_CP_PriorityOptions"),
    )
    for column, title, options, name in option_groups:
        ws.cell(1, column, title)
        for row_number, option in enumerate(options, start=2):
            ws.cell(row_number, column, option)
        letter = get_column_letter(column)
        ws.column_dimensions[letter].hidden = True
        _add_defined_name(
            workbook,
            name,
            f"'Methodology'!${letter}$2:${letter}${len(options) + 1}",
        )


def _add_content_plan_validations_and_colors(ws: Any, data_count: int) -> None:
    max_row = max(102, data_count + 101)
    validations = (
        ("D", "=_CP_FunnelOptions", "Choose TOFU, MOFU, or BOFU"),
        ("E", "=_CP_ContentTypeOptions", "Choose an approved content type"),
        ("N", "=_CP_ActionOptions", "Choose New, Existing, Update, or Consolidate"),
        ("P", "=_CP_PriorityOptions", "Choose P1, P2, or P3"),
    )
    for column, formula, prompt in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        validation.error = "Select a value from the approved list."
        validation.errorTitle = "Invalid value"
        validation.prompt = prompt
        validation.promptTitle = "Content Planner"
        validation.showErrorMessage = True
        validation.showInputMessage = True
        ws.add_data_validation(validation)
        validation.add(f"{column}2:{column}{max_row}")

    colors = {
        "TOFU": "BDD7EE",
        "MOFU": "C6E0B4",
        "BOFU": "FFE699",
        "New": "BDD7EE",
        "Existing": "C6E0B4",
        "Update": "FFE699",
        "Consolidate": "F4B7B2",
        "P1": "F4B7B2",
        "P2": "BDD7EE",
        "P3": "C6E0B4",
    }
    for value in FUNNEL_OPTIONS:
        ws.conditional_formatting.add(
            f"D2:D{max_row}",
            FormulaRule(formula=[f'$D2="{value}"'], fill=PatternFill("solid", fgColor=colors[value])),
        )
    for value in ACTION_OPTIONS:
        ws.conditional_formatting.add(
            f"N2:N{max_row}",
            FormulaRule(formula=[f'$N2="{value}"'], fill=PatternFill("solid", fgColor=colors[value])),
        )
    for value in PRIORITY_OPTIONS:
        ws.conditional_formatting.add(
            f"P2:P{max_row}",
            FormulaRule(formula=[f'$P2="{value}"'], fill=PatternFill("solid", fgColor=colors[value])),
        )

    content_type_colors = (
        ('LEFT($E2,4)="Tool"', "F4B183"),
        ('LEFT($E2,12)="Landing Page"', "9DC3E6"),
        ('LEFT($E2,8)="Template"', "A9D18E"),
        ('LEFT($E2,4)="Blog"', "C6E0B4"),
        ('OR(LEFT($E2,8)="Research",LEFT($E2,4)="Case")', "D9E1F2"),
    )
    for formula, color in content_type_colors:
        ws.conditional_formatting.add(
            f"E2:E{max_row}",
            FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=color)),
        )
    ws.conditional_formatting.add(
        f"O2:O{max_row}",
        CellIsRule(operator="greaterThanOrEqual", formula=["75"], fill=PatternFill("solid", fgColor="F4B7B2")),
    )
    ws.conditional_formatting.add(
        f"O2:O{max_row}",
        CellIsRule(operator="between", formula=["50", "74.9999"], fill=PatternFill("solid", fgColor="BDD7EE")),
    )
    ws.conditional_formatting.add(
        f"O2:O{max_row}",
        CellIsRule(operator="lessThan", formula=["50"], fill=PatternFill("solid", fgColor="C6E0B4")),
    )


def _add_provisional_score_comments(ws: Any, rows: Sequence[Mapping[str, Any]]) -> None:
    for row_number, row in enumerate(rows, start=2):
        if not row.get("score_is_provisional"):
            continue
        breakdown = row.get("score_breakdown")
        if not isinstance(breakdown, Mapping):
            breakdown = {}
        missing: list[str] = []
        if breakdown.get("demand") is None:
            missing.append("Demand")
        if breakdown.get("feasibility") is None:
            missing.append("Feasibility")
        lines = [
            "Provisional priority score; capped at 74 until missing Ahrefs evidence is restored.",
            "Missing component(s): " + (", ".join(missing) if missing else "see Methodology"),
        ]
        known_points = breakdown.get("known_points")
        known_maximum = breakdown.get("known_maximum")
        uncapped = breakdown.get("uncapped_provisional_score")
        if known_points is not None and known_maximum is not None:
            lines.append(f"Known points: {known_points}/{known_maximum}")
        if uncapped is not None:
            lines.append(f"Uncapped provisional score: {uncapped}")
        ws.cell(row=row_number, column=15).comment = Comment("\n".join(lines), "Span")


def _format_metrics(workbook: Any) -> None:
    plan = workbook["Content Plan"]
    for row in range(2, plan.max_row + 1):
        for column in (6, 7, 8, 15):
            plan.cell(row, column).number_format = "0"
        plan.cell(row, 10).number_format = "0.00"
    raw = workbook["Raw Keywords"]
    for row in range(2, raw.max_row + 1):
        for column in (4, 5, 6):
            raw.cell(row, column).number_format = "0"
        raw.cell(row, 8).number_format = "0.00"
    roadmap = workbook["Roadmap"]
    for row in range(2, roadmap.max_row + 1):
        roadmap.cell(row, 8).number_format = "0"


def _add_url_hyperlinks(ws: Any, column: int, metadata: Mapping[str, Any]) -> None:
    base_url = str(metadata.get("site_url") or "").rstrip("/")
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row, column)
        if not isinstance(cell.value, str) or not cell.value:
            continue
        target = cell.value
        if target.startswith(("http://", "https://")):
            pass
        elif target.startswith("/"):
            if base_url.startswith(("http://", "https://")):
                target = base_url + target
        else:
            continue
        cell.hyperlink = target
        cell.style = "Hyperlink"
        cell.alignment = Alignment(vertical="top", wrap_text=True)


def create_workbook(artifact: Mapping[str, Any]) -> Any:
    require_openpyxl()
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in SHEET_ORDER:
        workbook.create_sheet(sheet_name)

    metadata = artifact["metadata"]
    plan_headers = content_plan_headers(metadata)
    raw_headers = raw_keyword_headers(metadata)
    methodology_rows = _methodology_rows(artifact)

    _write_sheet(
        workbook["Content Plan"],
        plan_headers,
        artifact["content_plan"],
        _content_plan_values,
        "ContentPlanTable",
    )
    _write_sheet(
        workbook["Raw Keywords"],
        raw_headers,
        artifact["raw_keywords"],
        _raw_keyword_values,
        "RawKeywordsTable",
    )
    _write_sheet(
        workbook["Topic Map"],
        TOPIC_MAP_HEADERS,
        artifact["topic_map"],
        _topic_map_values,
        "TopicMapTable",
    )
    _write_sheet(
        workbook["Roadmap"],
        ROADMAP_HEADERS,
        artifact["roadmap"],
        _roadmap_values,
        "RoadmapTable",
    )
    _write_sheet(
        workbook["Strategy Notes"],
        STRATEGY_NOTES_HEADERS,
        artifact["strategy_notes"],
        _strategy_values,
        "StrategyNotesTable",
    )
    _write_sheet(
        workbook["Methodology"],
        METHODOLOGY_HEADERS,
        methodology_rows,
        _methodology_values,
        "MethodologyTable",
    )

    _add_validation_options(workbook)
    _add_content_plan_validations_and_colors(
        workbook["Content Plan"], len(artifact["content_plan"])
    )
    _add_provisional_score_comments(workbook["Content Plan"], artifact["content_plan"])
    _format_metrics(workbook)
    _add_url_hyperlinks(workbook["Content Plan"], 13, metadata)
    _add_url_hyperlinks(workbook["Topic Map"], 6, metadata)
    workbook.active = 0

    props = workbook.properties
    props.creator = "Span"
    props.lastModifiedBy = "Span"
    props.title = f"{metadata['project_name']} Content Plan - {metadata['country']['code']}"
    props.subject = (
        f"SEO/GEO whole-site plan for {metadata['country']['name']} "
        f"in {metadata['language']['name']}"
    )
    props.description = (
        f"Independent content plan for {metadata['country']['name']} "
        f"in {metadata['language']['name']}; keyword facts and metrics from Ahrefs MCP."
    )
    props.keywords = (
        f"schema={SCHEMA_VERSION};country={metadata['country']['code']};"
        f"language={metadata['language']['code']};source=Ahrefs MCP"
    )
    props.category = "SEO/GEO Content Strategy"
    props.language = metadata["language"]["code"]
    try:
        props.created = datetime.fromisoformat(metadata["generated_at"].replace("Z", "+00:00"))
    except ValueError:
        props.created = datetime.now(timezone.utc)
    props.modified = datetime.now(timezone.utc)
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.calculation.calcMode = "manual"
    return workbook


def build_xlsx(input_path: Path, output_path: Path, *, overwrite: bool = False) -> dict[str, Any]:
    require_openpyxl()
    if output_path.suffix.lower() != ".xlsx":
        raise ArtifactError("output must use the .xlsx extension; macro-enabled files are not supported")
    if output_path.exists() and not overwrite:
        raise ArtifactError(f"refusing to overwrite existing workbook: {output_path}")
    artifact = load_json_artifact(input_path, complete=True)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = create_workbook(artifact)
    workbook.save(output_path)
    return {
        "status": "ok",
        "schema_version": SCHEMA_VERSION,
        "output": str(output_path.resolve()),
        "country": artifact["metadata"]["country"],
        "language": artifact["metadata"]["language"],
        "counts": {
            "content_plan": len(artifact["content_plan"]),
            "raw_keywords": len(artifact["raw_keywords"]),
            "topic_map": len(artifact["topic_map"]),
            "roadmap": len(artifact["roadmap"]),
        },
    }


def _checkpoint_metadata(metadata: Mapping[str, Any]) -> dict[str, Any]:
    """Whitelist public metadata so credentials can never enter checkpoints."""
    return {
        "project_name": metadata["project_name"],
        "project_type": metadata["project_type"],
        "site_url": metadata["site_url"],
        "country": dict(metadata["country"]),
        "language": dict(metadata["language"]),
        "data_source": metadata["data_source"],
        "generated_at": metadata["generated_at"],
        "data_date": metadata["data_date"],
        "run_id": metadata["run_id"],
        "checkpoint_count": metadata["checkpoint_count"],
        "validation_result": metadata["validation_result"],
        "data_status": metadata["data_status"],
        "frontier_status": metadata["frontier_status"],
        "ahrefs_tools_used": list(metadata["ahrefs_tools_used"]),
        "frontier_items": [dict(item) for item in metadata["frontier_items"]],
        "ahrefs_units_before": metadata["ahrefs_units_before"],
        "ahrefs_units_after": metadata["ahrefs_units_after"],
        "site_domain_rating": metadata["site_domain_rating"],
    }


def write_checkpoints(input_path: Path, output_dir: Path) -> dict[str, Any]:
    artifact = load_json_artifact(input_path, complete=False)
    raw_rows = artifact["raw_keywords"]
    if not raw_rows:
        raise ArtifactError("raw_keywords must contain at least one row for checkpointing")
    output_dir.mkdir(parents=True, exist_ok=True)
    filename_pattern = re.compile(r"^raw-keywords-(\d{6})-(\d{6})\.json$")
    existing_ranges: list[tuple[int, int, Path]] = []
    for path in output_dir.iterdir():
        match = filename_pattern.fullmatch(path.name)
        if path.is_file() and match:
            existing_ranges.append((int(match.group(1)), int(match.group(2)), path))
    existing_ranges.sort()

    expected_start = 1
    existing_count = 0
    for expected_index, (start_number, end_number, path) in enumerate(
        existing_ranges, start=1
    ):
        if start_number != expected_start:
            raise ArtifactError(
                f"checkpoint sequence is not a contiguous prefix at {path}"
            )
        if end_number < start_number or end_number - start_number + 1 > CHECKPOINT_BATCH_SIZE:
            raise ArtifactError(f"existing checkpoint has an invalid row range: {path}")
        if end_number > len(raw_rows):
            raise ArtifactError(
                f"existing checkpoint extends beyond the current artifact: {path}"
            )
        rows = raw_rows[start_number - 1 : end_number]
        try:
            existing_payload = json.loads(path.read_text(encoding="utf-8-sig"))
        except (OSError, json.JSONDecodeError) as exc:
            raise ArtifactError(f"cannot validate existing checkpoint {path}: {exc}") from exc
        existing_rows = existing_payload.get("raw_keywords")
        existing_meta = existing_payload.get("metadata")
        if not isinstance(existing_meta, Mapping):
            raise ArtifactError(f"existing checkpoint has invalid metadata: {path}")
        if existing_payload.get("schema_version") != SCHEMA_VERSION:
            raise ArtifactError(f"existing checkpoint has an incompatible schema version: {path}")
        if existing_payload.get("checkpoint_version") != CHECKPOINT_VERSION:
            raise ArtifactError(f"existing checkpoint has an incompatible checkpoint version: {path}")
        current_meta = artifact["metadata"]
        identity_fields = (
            "project_name",
            "project_type",
            "site_url",
            "country",
            "language",
            "data_source",
            "run_id",
        )
        mismatched_identity = [
            field
            for field in identity_fields
            if existing_meta.get(field) != current_meta.get(field)
        ]
        if mismatched_identity:
            raise ArtifactError(
                f"existing checkpoint belongs to a different run ({', '.join(mismatched_identity)}): {path}"
            )
        if existing_meta.get("checkpoint_index") != expected_index:
            raise ArtifactError(f"existing checkpoint has a non-contiguous checkpoint_index: {path}")
        existing_frontier = existing_meta.get("frontier_items")
        if not isinstance(existing_frontier, list) or not existing_frontier:
            raise ArtifactError(f"existing checkpoint is missing frontier identity: {path}")
        current_frontier = {
            item["frontier_id"]: item for item in current_meta["frontier_items"]
        }
        frontier_identity_fields = (
            "frontier_id",
            "source_tool",
            "target",
            "mode",
            "filters",
            "selected_fields",
        )
        allowed_status_transitions = {
            "Queued": {"Queued", "Partial", "Completed", "Exhausted", "Failed"},
            "Partial": {"Partial", "Completed", "Exhausted", "Failed"},
            "Completed": {"Completed", "Exhausted"},
            "Exhausted": {"Exhausted"},
            "Failed": {"Failed"},
        }
        for frontier_item in existing_frontier:
            if not isinstance(frontier_item, Mapping):
                raise ArtifactError(f"existing checkpoint has invalid frontier data: {path}")
            frontier_id = frontier_item.get("frontier_id")
            if not isinstance(frontier_id, str) or not frontier_id:
                raise ArtifactError(f"existing checkpoint has invalid frontier identity: {path}")
            current_item = current_frontier.get(frontier_id)
            if current_item is None or any(
                current_item.get(field) != frontier_item.get(field)
                for field in frontier_identity_fields
            ):
                raise ArtifactError(
                    f"existing checkpoint frontier does not match the current run: {path}"
                )
            previous_status = frontier_item.get("status")
            current_status = current_item.get("status")
            if current_status not in allowed_status_transitions.get(previous_status, set()):
                raise ArtifactError(
                    f"existing checkpoint frontier status regressed for {frontier_id!r}: {path}"
                )
        existing_tools = existing_meta.get("ahrefs_tools_used")
        if (
            not isinstance(existing_tools, list)
            or any(not isinstance(tool, str) for tool in existing_tools)
            or not set(existing_tools).issubset(set(current_meta["ahrefs_tools_used"]))
        ):
            raise ArtifactError(f"existing checkpoint tool identity is incompatible: {path}")
        batch_stats = existing_meta.get("batch_stats")
        expected_source_tools = sorted(
            {source for row in rows for source in row["source_tool"]},
            key=str.casefold,
        )
        expected_decisions = {
            decision: sum(1 for row in rows if row["decision"] == decision)
            for decision in ("Include", "Exclude", "Defer")
        }
        if (
            not isinstance(batch_stats, Mapping)
            or batch_stats.get("persisted_rows") != len(rows)
            or batch_stats.get("source_tools") != expected_source_tools
            or batch_stats.get("included_rows") != expected_decisions["Include"]
            or batch_stats.get("excluded_rows") != expected_decisions["Exclude"]
            or batch_stats.get("deferred_rows") != expected_decisions["Defer"]
        ):
            raise ArtifactError(f"existing checkpoint has invalid batch statistics: {path}")
        if (
            existing_rows != rows
            or existing_meta.get("row_start") != start_number
            or existing_meta.get("row_end") != end_number
            or existing_meta.get("country") != artifact["metadata"]["country"]
            or existing_meta.get("language") != artifact["metadata"]["language"]
        ):
            raise ArtifactError(
                f"existing checkpoint does not match the current artifact; refusing to overwrite: {path}"
            )
        existing_count += 1
        expected_start = end_number + 1

    completed_rows = expected_start - 1
    if completed_rows == len(raw_rows):
        raise ArtifactError(
            "refusing to overwrite existing checkpoint(s): "
            + ", ".join(str(path) for _, _, path in existing_ranges)
        )

    targets: list[tuple[Path, int, int, list[dict[str, Any]]]] = []
    for start in range(completed_rows, len(raw_rows), CHECKPOINT_BATCH_SIZE):
        end = min(start + CHECKPOINT_BATCH_SIZE, len(raw_rows))
        target = output_dir / f"raw-keywords-{start + 1:06d}-{end:06d}.json"
        if target.exists():
            raise ArtifactError(f"refusing to overwrite existing checkpoint: {target}")
        targets.append((target, start, end, raw_rows[start:end]))

    created: list[str] = []
    for batch_index, (target, start, end, rows) in enumerate(
        targets, start=existing_count + 1
    ):
        decision_counts = {
            decision: sum(1 for row in rows if row["decision"] == decision)
            for decision in ("Include", "Exclude", "Defer")
        }
        batch_sources = sorted(
            {source for row in rows for source in row["source_tool"]},
            key=str.casefold,
        )
        payload = {
            "schema_version": SCHEMA_VERSION,
            "checkpoint_version": CHECKPOINT_VERSION,
            "metadata": {
                **_checkpoint_metadata(artifact["metadata"]),
                "checkpoint_index": batch_index,
                "row_start": start + 1,
                "row_end": end,
                "raw_keyword_total": len(raw_rows),
                "batch_size": CHECKPOINT_BATCH_SIZE,
                "batch_stats": {
                    "persisted_rows": len(rows),
                    "included_rows": decision_counts["Include"],
                    "excluded_rows": decision_counts["Exclude"],
                    "deferred_rows": decision_counts["Defer"],
                    "source_tools": batch_sources,
                },
            },
            "raw_keywords": rows,
        }
        with target.open("x", encoding="utf-8", newline="\n") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
        created.append(str(target.resolve()))
    return {
        "status": "ok",
        "schema_version": SCHEMA_VERSION,
        "checkpoint_version": CHECKPOINT_VERSION,
        "batch_size": CHECKPOINT_BATCH_SIZE,
        "raw_keywords": len(raw_rows),
        "existing_checkpoints": existing_count,
        "checkpoints": created,
    }


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Build or checkpoint a content-planner JSON artifact without network access."
    )
    subparsers = parser.add_subparsers(dest="command", required=True)

    build = subparsers.add_parser("build", help="render a validated JSON artifact to .xlsx")
    build.add_argument("--input", required=True, type=Path, help="UTF-8 JSON artifact")
    build.add_argument("--output", required=True, type=Path, help="destination .xlsx file")
    build.add_argument(
        "--overwrite",
        action="store_true",
        help="explicitly allow replacing an existing workbook",
    )
    build.add_argument(
        "--checkpoint-dir",
        type=Path,
        help="also write non-overwriting 100-row raw-keyword checkpoints",
    )

    checkpoint = subparsers.add_parser(
        "checkpoint", help="write one non-overwriting checkpoint per 100 raw keywords"
    )
    checkpoint.add_argument("--input", required=True, type=Path, help="UTF-8 JSON artifact")
    checkpoint.add_argument("--output-dir", required=True, type=Path)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = _parser()
    args = parser.parse_args(argv)
    try:
        if args.command == "build":
            if args.output.suffix.lower() != ".xlsx":
                raise ArtifactError(
                    "output must use the .xlsx extension; macro-enabled files are not supported"
                )
            if args.output.exists() and not args.overwrite:
                raise ArtifactError(f"refusing to overwrite existing workbook: {args.output}")
            if args.checkpoint_dir is not None:
                # Validate the complete artifact before writing any checkpoint side effects.
                load_json_artifact(args.input, complete=True)
                checkpoint_result = write_checkpoints(args.input, args.checkpoint_dir)
            else:
                checkpoint_result = None
            result = build_xlsx(args.input, args.output, overwrite=args.overwrite)
            if checkpoint_result is not None:
                result["checkpoints"] = checkpoint_result["checkpoints"]
        else:
            result = write_checkpoints(args.input, args.output_dir)
    except (ArtifactError, RuntimeError, OSError) as exc:
        print(json.dumps({"status": "error", "error": str(exc)}, ensure_ascii=False), file=sys.stderr)
        return 2
    print(json.dumps(result, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
