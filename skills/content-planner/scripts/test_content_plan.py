#!/usr/bin/env python3
"""Offline end-to-end tests for the content-planner workbook scripts."""

from __future__ import annotations

import json
import os
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path
from typing import Any


OPENPYXL_IMPORT_ERROR: ModuleNotFoundError | None = None
try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - caller environment dependent
    OPENPYXL_IMPORT_ERROR = exc


SCRIPT_DIR = Path(__file__).resolve().parent
BUILDER = SCRIPT_DIR / "build_content_plan_xlsx.py"
VALIDATOR = SCRIPT_DIR / "validate_content_plan.py"


def fixture_artifact() -> dict[str, Any]:
    formula_keyword = '=HYPERLINK("https://evil.invalid","危険")'
    content_plan = [
        {
            "primary_keyword": "AI音楽ジェネレーター",
            "supporting_keywords": ["AI 作曲ツール"],
            "topic": "AI 音楽作成",
            "funnel": "BOFU",
            "content_type": "Tool",
            "kd": 0,
            "volume": 1200,
            "traffic_potential": "N/A",
            "search_intent": ["Informational", "Commercial"],
            "cpc_usd": 0,
            "parent_topic": "AI 音楽",
            "serp_features": ["AI Overview", "Video"],
            "url": "/ai-music-generator",
            "action": "Existing",
            "priority_score": 82,
            "priority": "P1",
            "commercial_relevance_score": 30,
            "strategic_value_score": 10,
            "score_breakdown": {"funnel": 25, "demand": 17},
        },
        {
            "primary_keyword": formula_keyword,
            "supporting_keywords": [],
            "topic": "安全测试",
            "funnel": "TOFU",
            "content_type": "Blog - Info",
            "kd": "N/A",
            "volume": 0,
            "traffic_potential": 0,
            "search_intent": ["Informational"],
            "cpc_usd": "N/A",
            "parent_topic": "N/A",
            "serp_features": [],
            "url": "/formula-safety-test",
            "action": "New",
            "priority_score": 45,
            "priority": "P3",
            "score_is_provisional": True,
            "score_breakdown": {
                "demand": 0,
                "feasibility": None,
                "known_points": 38,
                "known_maximum": 85,
                "uncapped_provisional_score": 45
            },
        },
        {
            "primary_keyword": "AI 作曲 方法",
            "supporting_keywords": [],
            "topic": "AI 音楽学習",
            "funnel": "MOFU",
            "content_type": "Blog - How to",
            "kd": 12,
            "volume": 100,
            "traffic_potential": 250,
            "search_intent": ["Informational"],
            "cpc_usd": 1.25,
            "parent_topic": "AI 作曲",
            "serp_features": ["People Also Ask"],
            "url": "https://example.jp/blog/ai-compose-guide",
            "action": "New",
            "priority_score": 50,
            "priority": "P2",
        },
    ]
    raw_keywords = [
        {
            "keyword": "AI音楽ジェネレーター",
            "country": "JP",
            "language": "ja",
            "volume": 1200,
            "kd": 0,
            "traffic_potential": "N/A",
            "search_intent": ["Informational", "Commercial"],
            "cpc_usd": 0,
            "parent_topic": "AI 音楽",
            "serp_features": ["AI Overview", "Video"],
            "source_tool": "keywords-explorer-matching-terms",
            "seed_or_competitor": "AI 音楽",
            "serp_updated": "2026-07-14",
            "decision": "Include",
            "decision_reason": "High product fit",
            "mapped_primary_keyword": "AI音楽ジェネレーター",
            "needs_review": False,
        },
        {
            "keyword": "AI 作曲ツール",
            "country": "Japan",
            "language": "Japanese",
            "volume": 250,
            "kd": 4,
            "traffic_potential": 500,
            "search_intent": ["Commercial"],
            "cpc_usd": 0.75,
            "parent_topic": "AI 音楽",
            "serp_features": [],
            "source_tool": "keywords-explorer-related-terms",
            "seed_or_competitor": "AI 音楽",
            "serp_updated": "2026-07-14",
            "decision": "Include",
            "decision_reason": "Same SERP cluster",
            "mapped_primary_keyword": "AI音楽ジェネレーター",
            "needs_review": "No",
        },
        {
            "keyword": formula_keyword,
            "country": "JP",
            "language": "ja",
            "volume": 0,
            "kd": "N/A",
            "traffic_potential": 0,
            "search_intent": ["Informational"],
            "cpc_usd": "N/A",
            "parent_topic": "N/A",
            "serp_features": [],
            "source_tool": "keywords-explorer-matching-terms",
            "seed_or_competitor": "安全テスト",
            "serp_updated": "2026-07-14",
            "decision": "Include",
            "decision_reason": "+formula-like audit text",
            "mapped_primary_keyword": formula_keyword,
            "needs_review": True,
        },
        {
            "keyword": "AI 作曲 方法",
            "country": "JP",
            "language": "ja",
            "volume": 100,
            "kd": 12,
            "traffic_potential": 250,
            "search_intent": ["Informational"],
            "cpc_usd": 1.25,
            "parent_topic": "AI 作曲",
            "serp_features": ["People Also Ask"],
            "source_tool": "keywords-explorer-matching-terms",
            "seed_or_competitor": "AI 作曲",
            "serp_updated": "2026-07-14",
            "decision": "Include",
            "decision_reason": "Supports journey",
            "mapped_primary_keyword": "AI 作曲 方法",
            "needs_review": False,
        },
    ]
    for number in range(1, 101):
        raw_keywords.append(
            {
                "keyword": f"追加キーワード {number}",
                "country": "JP",
                "language": "ja",
                "volume": 0 if number % 2 else "N/A",
                "kd": "N/A",
                "traffic_potential": "N/A",
                "search_intent": ["Informational"],
                "cpc_usd": "N/A",
                "parent_topic": "AI 音楽",
                "serp_features": [],
                "source_tool": "keywords-explorer-matching-terms",
                "seed_or_competitor": "AI 音楽",
                "serp_updated": "2026-07-14",
                "decision": "Exclude",
                "decision_reason": "Low relevance",
                "mapped_primary_keyword": "",
                "needs_review": number == 100,
            }
        )

    topic_map = [
        {
            "topic": row["topic"],
            "page_level": "1" if index == 0 else "2",
            "primary_keyword": row["primary_keyword"],
            "page_role": "Pillar" if index == 0 else "Cluster",
            "parent_page": "" if index == 0 else "AI音楽ジェネレーター",
            "url": row["url"],
            "link_up_to": [] if index == 0 else ["AI音楽ジェネレーター"],
            "relevant_cross_links": ["AI 作曲 方法"] if index == 0 else [],
        }
        for index, row in enumerate(content_plan)
    ]
    roadmap = [
        {
            "phase": "Phase 1" if row["priority"] == "P1" else "Phase 2",
            "sequence": index,
            "primary_keyword": row["primary_keyword"],
            "topic": row["topic"],
            "funnel": row["funnel"],
            "content_type": row["content_type"],
            "action": row["action"],
            "priority_score": row["priority_score"],
            "priority": row["priority"],
            "dependency": [] if index == 1 else ["AI音楽ジェネレーター"],
            "internal_link_targets": ["AI音楽ジェネレーター"],
            "reason": "满足用户搜索意图，并建立主题权威。",
        }
        for index, row in enumerate(content_plan, start=1)
    ]
    return {
        "schema_version": "1.0",
        "metadata": {
            "project_name": "多语言 AI 音乐站",
            "project_type": "Existing",
            "site_url": "https://example.jp",
            "country": {"name": "Japan", "code": "JP", "volume_label": "JP"},
            "language": {"name": "Japanese", "code": "ja"},
            "data_source": "Ahrefs MCP",
            "generated_at": "2026-07-15T10:00:00+09:00",
            "data_date": "2026-07-14",
            "run_id": "jp-20260715-01",
            "checkpoint_count": 2,
            "validation_result": "Pending",
            "data_status": "Partial",
            "frontier_status": "Stopped by user after 104 rows",
            "ahrefs_tools_used": [
                "keywords-explorer-matching-terms",
                "keywords-explorer-related-terms",
                "serp-overview-serp-overview",
            ],
            "frontier_items": [
                {
                    "frontier_id": "matching-ai-music-jp-001",
                    "source_tool": "keywords-explorer-matching-terms",
                    "target": "AI 音楽",
                    "mode": "terms",
                    "filters": {"country": "JP", "language": "ja"},
                    "selected_fields": ["keyword", "volume", "difficulty", "cpc"],
                    "status": "Completed",
                    "returned_rows": 103,
                    "new_unique_rows": 103,
                    "duplicate_rows": 0,
                    "included_rows": 3,
                    "excluded_rows": 100,
                    "units_before": 5000,
                    "units_after": 4500,
                    "collected_at": "2026-07-14T10:00:00+09:00"
                },
                {
                    "frontier_id": "related-ai-music-jp-001",
                    "source_tool": "keywords-explorer-related-terms",
                    "target": "AI 音楽",
                    "mode": "also-talk-about",
                    "filters": {"country": "JP", "language": "ja"},
                    "selected_fields": ["keyword", "volume", "difficulty", "cpc"],
                    "status": "Completed",
                    "returned_rows": 1,
                    "new_unique_rows": 1,
                    "duplicate_rows": 0,
                    "included_rows": 1,
                    "excluded_rows": 0,
                    "units_before": 4500,
                    "units_after": 4400,
                    "collected_at": "2026-07-14T10:05:00+09:00"
                }
            ],
            "ahrefs_units_before": 5000,
            "ahrefs_units_after": 4500,
            "site_domain_rating": 23,
        },
        "content_plan": content_plan,
        "raw_keywords": raw_keywords,
        "topic_map": topic_map,
        "roadmap": roadmap,
        "strategy_notes": [
            {
                "section": "Content Marketing Mission",
                "item": "核心使命",
                "details": "日本のクリエイターが AI 音楽を安全に制作するための情報を提供する。",
            },
            {
                "section": "SEO/GEO",
                "item": "AI Overview",
                "details": "覆盖定义、步骤、限制与常见问题，并保留 Ahrefs 的 SERP 证据。",
            },
        ],
        "methodology": [
            {
                "field": "Priority Scoring",
                "value": "30/25/20/15/10",
                "notes": "Commercial relevance, funnel, demand, feasibility, strategy.",
            },
            {
                "field": "SERP Overlap",
                "value": "5+ merge; 0-2 split; 3-4 review",
                "notes": "Only Ahrefs MCP SERP rows are used.",
            },
            {
                "field": "Missing Metrics",
                "value": "N/A",
                "notes": "Never replace unavailable metrics with zero.",
            },
        ],
    }


class ContentPlanScriptsTest(unittest.TestCase):
    def setUp(self) -> None:
        if OPENPYXL_IMPORT_ERROR is not None:
            self.fail(
                "Missing dependency 'openpyxl'. Install it in the active Python environment "
                "before running these tests."
            )
        self.temp_context = tempfile.TemporaryDirectory()
        self.temp_dir = Path(self.temp_context.name)
        self.input_json = self.temp_dir / "artifact.json"
        self.output_xlsx = self.temp_dir / "content-plan-jp.xlsx"
        self.checkpoint_dir = self.temp_dir / "checkpoints"
        self.artifact = fixture_artifact()
        self.input_json.write_text(
            json.dumps(self.artifact, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )

    def tearDown(self) -> None:
        self.temp_context.cleanup()

    def run_cli(self, *arguments: object) -> subprocess.CompletedProcess[str]:
        environment = dict(os.environ)
        environment["PYTHONDONTWRITEBYTECODE"] = "1"
        return subprocess.run(
            [sys.executable, *(str(argument) for argument in arguments)],
            text=True,
            capture_output=True,
            check=False,
            env=environment,
        )

    def build(self) -> subprocess.CompletedProcess[str]:
        return self.run_cli(
            BUILDER,
            "build",
            "--input",
            self.input_json,
            "--output",
            self.output_xlsx,
            "--checkpoint-dir",
            self.checkpoint_dir,
        )

    def test_build_validate_multilingual_safety_and_checkpoints(self) -> None:
        built = self.build()
        self.assertEqual(built.returncode, 0, built.stderr)
        build_result = json.loads(built.stdout)
        self.assertEqual(build_result["counts"]["raw_keywords"], 104)
        checkpoints = sorted(self.checkpoint_dir.glob("*.json"))
        self.assertEqual(len(checkpoints), 2)
        first = json.loads(checkpoints[0].read_text(encoding="utf-8"))
        second = json.loads(checkpoints[1].read_text(encoding="utf-8"))
        self.assertEqual(len(first["raw_keywords"]), 100)
        self.assertEqual(len(second["raw_keywords"]), 4)
        self.assertEqual(first["metadata"]["checkpoint_index"], 1)
        self.assertEqual(second["metadata"]["checkpoint_index"], 2)
        self.assertEqual(first["metadata"]["batch_stats"]["persisted_rows"], 100)
        self.assertEqual(len(first["metadata"]["frontier_items"]), 2)
        serialized_checkpoints = json.dumps([first, second], ensure_ascii=False).casefold()
        self.assertNotIn("ahrefs_mcp_key", serialized_checkpoints)
        self.assertNotIn("api_key", serialized_checkpoints)

        validated = self.run_cli(
            VALIDATOR,
            "--workbook",
            self.output_xlsx,
            "--source-json",
            self.input_json,
            "--mark-valid",
        )
        self.assertEqual(validated.returncode, 0, validated.stdout + validated.stderr)
        validation_result = json.loads(validated.stdout)
        self.assertTrue(validation_result["valid"])
        self.assertTrue(validation_result["validation_result_marked"])
        self.assertEqual(validation_result["metadata"]["Validation Result"], "Valid")
        self.assertEqual(validation_result["metadata"]["Country Code"], "JP")
        self.assertEqual(validation_result["metadata"]["Language Code"], "ja")

        workbook = load_workbook(self.output_xlsx, data_only=False)
        self.assertEqual(
            workbook.sheetnames,
            [
                "Content Plan",
                "Raw Keywords",
                "Topic Map",
                "Roadmap",
                "Strategy Notes",
                "Methodology",
            ],
        )
        self.assertEqual(workbook.active.title, "Content Plan")
        self.assertEqual(workbook["Content Plan"]["G1"].value, "Volume (JP)")
        validation_rows = {
            row[0].value: row[1].value
            for row in workbook["Methodology"].iter_rows(min_row=2, max_col=2)
            if row[0].value
        }
        self.assertEqual(validation_rows["Validation Result"], "Valid")

        formula_rows = [
            row
            for row in workbook["Content Plan"].iter_rows(min_row=2, max_col=16)
            if isinstance(row[0].value, str) and "HYPERLINK" in row[0].value
        ]
        self.assertEqual(len(formula_rows), 1)
        self.assertTrue(formula_rows[0][0].value.startswith("'="))
        self.assertNotEqual(formula_rows[0][0].data_type, "f")
        self.assertEqual(formula_rows[0][5].value, "N/A")
        self.assertEqual(formula_rows[0][6].value, 0)
        self.assertEqual(formula_rows[0][7].value, 0)
        self.assertEqual(formula_rows[0][9].value, "N/A")
        self.assertIsNotNone(formula_rows[0][14].comment)
        self.assertIn("Provisional priority score", formula_rows[0][14].comment.text)

        for row in workbook["Content Plan"].iter_rows(min_row=2, max_col=16):
            if row[0].value == "AI音楽ジェネレーター":
                self.assertEqual(row[5].value, 0)
                self.assertEqual(row[7].value, "N/A")
                self.assertEqual(row[9].value, 0)
                self.assertIsNotNone(row[12].hyperlink)
                break
        else:
            self.fail("multilingual primary keyword was not preserved")

    def test_non_overwrite_and_validator_detects_priority_tampering(self) -> None:
        built = self.build()
        self.assertEqual(built.returncode, 0, built.stderr)

        refused = self.run_cli(
            BUILDER,
            "build",
            "--input",
            self.input_json,
            "--output",
            self.output_xlsx,
        )
        self.assertEqual(refused.returncode, 2)
        self.assertIn("refusing to overwrite", refused.stderr)

        overwritten = self.run_cli(
            BUILDER,
            "build",
            "--input",
            self.input_json,
            "--output",
            self.output_xlsx,
            "--overwrite",
        )
        self.assertEqual(overwritten.returncode, 0, overwritten.stderr)

        checkpoint_refused = self.run_cli(
            BUILDER,
            "checkpoint",
            "--input",
            self.input_json,
            "--output-dir",
            self.checkpoint_dir,
        )
        self.assertEqual(checkpoint_refused.returncode, 2)
        self.assertIn("refusing to overwrite existing checkpoint", checkpoint_refused.stderr)

        workbook = load_workbook(self.output_xlsx)
        sheet = workbook["Content Plan"]
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 16).value == "P1":
                sheet.cell(row, 15, 74)
                break
        tampered = self.temp_dir / "tampered.xlsx"
        workbook.save(tampered)
        validation = self.run_cli(VALIDATOR, "--workbook", tampered)
        self.assertEqual(validation.returncode, 1, validation.stdout + validation.stderr)
        report = json.loads(validation.stdout)
        self.assertFalse(report["valid"])
        self.assertTrue(any("requires P2" in error for error in report["errors"]))

    def test_checkpoint_writer_appends_only_new_batches(self) -> None:
        first_hundred = fixture_artifact()
        first_hundred["raw_keywords"] = first_hundred["raw_keywords"][:100]
        initial_json = self.temp_dir / "first-100.json"
        initial_json.write_text(
            json.dumps(first_hundred, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        initial = self.run_cli(
            BUILDER,
            "checkpoint",
            "--input",
            initial_json,
            "--output-dir",
            self.checkpoint_dir,
        )
        self.assertEqual(initial.returncode, 0, initial.stderr)
        first_path = self.checkpoint_dir / "raw-keywords-000001-000100.json"
        self.assertTrue(first_path.exists())
        first_bytes = first_path.read_bytes()

        appended = self.run_cli(
            BUILDER,
            "checkpoint",
            "--input",
            self.input_json,
            "--output-dir",
            self.checkpoint_dir,
        )
        self.assertEqual(appended.returncode, 0, appended.stderr)
        result = json.loads(appended.stdout)
        self.assertEqual(result["existing_checkpoints"], 1)
        self.assertEqual(len(result["checkpoints"]), 1)
        self.assertTrue(
            (self.checkpoint_dir / "raw-keywords-000101-000104.json").exists()
        )
        appended_payload = json.loads(
            (self.checkpoint_dir / "raw-keywords-000101-000104.json").read_text(
                encoding="utf-8"
            )
        )
        self.assertEqual(appended_payload["metadata"]["checkpoint_index"], 2)
        self.assertEqual(first_path.read_bytes(), first_bytes)

    def test_checkpoint_writer_resumes_after_a_partial_batch(self) -> None:
        first_eighty = fixture_artifact()
        first_eighty["raw_keywords"] = first_eighty["raw_keywords"][:80]
        partial_json = self.temp_dir / "first-80.json"
        partial_json.write_text(
            json.dumps(first_eighty, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        initial = self.run_cli(
            BUILDER,
            "checkpoint",
            "--input",
            partial_json,
            "--output-dir",
            self.checkpoint_dir,
        )
        self.assertEqual(initial.returncode, 0, initial.stderr)

        resumed = self.run_cli(
            BUILDER,
            "checkpoint",
            "--input",
            self.input_json,
            "--output-dir",
            self.checkpoint_dir,
        )
        self.assertEqual(resumed.returncode, 0, resumed.stderr)
        result = json.loads(resumed.stdout)
        self.assertEqual(result["existing_checkpoints"], 1)
        self.assertEqual(len(result["checkpoints"]), 1)
        self.assertTrue(
            (self.checkpoint_dir / "raw-keywords-000081-000104.json").exists()
        )

    def test_builder_rejects_credential_shaped_artifact_fields(self) -> None:
        unsafe = fixture_artifact()
        unsafe["metadata"]["authorization"] = "Bearer " + "not-a-real-token-but-still-forbidden"
        unsafe_json = self.temp_dir / "unsafe.json"
        unsafe_json.write_text(
            json.dumps(unsafe, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        result = self.run_cli(
            BUILDER,
            "build",
            "--input",
            unsafe_json,
            "--output",
            self.output_xlsx,
        )
        self.assertEqual(result.returncode, 2)
        self.assertIn("credential-like field is forbidden", result.stderr)

    def test_builder_rejects_missing_or_undeclared_ahrefs_source(self) -> None:
        missing_source = fixture_artifact()
        del missing_source["raw_keywords"][0]["source_tool"]
        missing_json = self.temp_dir / "missing-source.json"
        missing_json.write_text(
            json.dumps(missing_source, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        missing = self.run_cli(
            BUILDER,
            "build",
            "--input",
            missing_json,
            "--output",
            self.output_xlsx,
        )
        self.assertEqual(missing.returncode, 2)
        self.assertIn("source_tool must name at least one Ahrefs MCP tool", missing.stderr)

        undeclared = fixture_artifact()
        undeclared["raw_keywords"][0]["source_tool"] = "unknown-ahrefs-tool"
        undeclared_json = self.temp_dir / "undeclared-source.json"
        undeclared_json.write_text(
            json.dumps(undeclared, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        result = self.run_cli(
            BUILDER,
            "build",
            "--input",
            undeclared_json,
            "--output",
            self.output_xlsx,
        )
        self.assertEqual(result.returncode, 2)
        self.assertIn("not listed in metadata.ahrefs_tools_used", result.stderr)

        missing_frontier = fixture_artifact()
        missing_frontier["metadata"]["frontier_items"] = [
            item
            for item in missing_frontier["metadata"]["frontier_items"]
            if item["source_tool"] != "keywords-explorer-related-terms"
        ]
        missing_frontier_json = self.temp_dir / "missing-frontier.json"
        missing_frontier_json.write_text(
            json.dumps(missing_frontier, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        result = self.run_cli(
            BUILDER,
            "build",
            "--input",
            missing_frontier_json,
            "--output",
            self.output_xlsx,
        )
        self.assertEqual(result.returncode, 2)
        self.assertIn("has no completed, partial, or exhausted frontier item", result.stderr)

    def test_nfkc_and_whitespace_keyword_duplicates_are_rejected(self) -> None:
        duplicate = fixture_artifact()
        duplicate_row = dict(duplicate["raw_keywords"][0])
        duplicate_row["keyword"] = "ＡＩ音楽ジェネレーター"
        duplicate["raw_keywords"].append(duplicate_row)
        duplicate_json = self.temp_dir / "nfkc-duplicate.json"
        duplicate_json.write_text(
            json.dumps(duplicate, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        result = self.run_cli(
            BUILDER,
            "build",
            "--input",
            duplicate_json,
            "--output",
            self.output_xlsx,
        )
        self.assertEqual(result.returncode, 2)
        self.assertIn("duplicates another raw keyword", result.stderr)

    def test_checkpoint_resume_rejects_another_project_identity(self) -> None:
        first_hundred = fixture_artifact()
        first_hundred["raw_keywords"] = first_hundred["raw_keywords"][:100]
        initial_json = self.temp_dir / "identity-first-100.json"
        initial_json.write_text(
            json.dumps(first_hundred, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        initial = self.run_cli(
            BUILDER,
            "checkpoint",
            "--input",
            initial_json,
            "--output-dir",
            self.checkpoint_dir,
        )
        self.assertEqual(initial.returncode, 0, initial.stderr)

        different_project = fixture_artifact()
        different_project["metadata"]["project_name"] = "Different project"
        different_json = self.temp_dir / "different-project.json"
        different_json.write_text(
            json.dumps(different_project, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        resumed = self.run_cli(
            BUILDER,
            "checkpoint",
            "--input",
            different_json,
            "--output-dir",
            self.checkpoint_dir,
        )
        self.assertEqual(resumed.returncode, 2)
        self.assertIn("belongs to a different run", resumed.stderr)

    def test_checkpoint_resume_allows_forward_frontier_status_progress(self) -> None:
        queued_item = {
            "frontier_id": "suggestions-future-jp-001",
            "source_tool": "keywords-explorer-search-suggestions",
            "target": "AI 音楽 作成",
            "mode": "suggestions",
            "filters": {"country": "JP", "language": "ja"},
            "selected_fields": ["keyword", "volume", "difficulty", "cpc"],
            "status": "Queued",
            "returned_rows": 0,
            "new_unique_rows": 0,
            "duplicate_rows": 0,
            "included_rows": 0,
            "excluded_rows": 0,
            "units_before": "N/A",
            "units_after": "N/A",
            "collected_at": ""
        }
        first_hundred = fixture_artifact()
        first_hundred["raw_keywords"] = first_hundred["raw_keywords"][:100]
        first_hundred["metadata"]["ahrefs_tools_used"].append(
            "keywords-explorer-search-suggestions"
        )
        first_hundred["metadata"]["frontier_items"].append(queued_item)
        first_json = self.temp_dir / "queued-first-100.json"
        first_json.write_text(
            json.dumps(first_hundred, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        initial = self.run_cli(
            BUILDER,
            "checkpoint",
            "--input",
            first_json,
            "--output-dir",
            self.checkpoint_dir,
        )
        self.assertEqual(initial.returncode, 0, initial.stderr)

        progressed = fixture_artifact()
        progressed["metadata"]["ahrefs_tools_used"].append(
            "keywords-explorer-search-suggestions"
        )
        completed_item = dict(queued_item)
        completed_item.update(
            {
                "status": "Completed",
                "returned_rows": 20,
                "new_unique_rows": 0,
                "units_before": 4400,
                "units_after": 4300,
                "collected_at": "2026-07-14T10:10:00+09:00",
            }
        )
        progressed["metadata"]["frontier_items"].append(completed_item)
        progressed_json = self.temp_dir / "frontier-progressed.json"
        progressed_json.write_text(
            json.dumps(progressed, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        resumed = self.run_cli(
            BUILDER,
            "checkpoint",
            "--input",
            progressed_json,
            "--output-dir",
            self.checkpoint_dir,
        )
        self.assertEqual(resumed.returncode, 0, resumed.stderr)
        self.assertTrue(
            (self.checkpoint_dir / "raw-keywords-000101-000104.json").exists()
        )

    def test_checkpoint_resume_rejects_tampered_batch_statistics(self) -> None:
        first_hundred = fixture_artifact()
        first_hundred["raw_keywords"] = first_hundred["raw_keywords"][:100]
        first_json = self.temp_dir / "stats-first-100.json"
        first_json.write_text(
            json.dumps(first_hundred, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        initial = self.run_cli(
            BUILDER,
            "checkpoint",
            "--input",
            first_json,
            "--output-dir",
            self.checkpoint_dir,
        )
        self.assertEqual(initial.returncode, 0, initial.stderr)
        checkpoint_path = self.checkpoint_dir / "raw-keywords-000001-000100.json"
        checkpoint = json.loads(checkpoint_path.read_text(encoding="utf-8"))
        checkpoint["metadata"]["batch_stats"]["included_rows"] += 1
        checkpoint_path.write_text(
            json.dumps(checkpoint, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        resumed = self.run_cli(
            BUILDER,
            "checkpoint",
            "--input",
            self.input_json,
            "--output-dir",
            self.checkpoint_dir,
        )
        self.assertEqual(resumed.returncode, 2)
        self.assertIn("invalid batch statistics", resumed.stderr)

    def test_validator_preflights_report_path_before_marking_workbook(self) -> None:
        built = self.build()
        self.assertEqual(built.returncode, 0, built.stderr)
        report_path = self.temp_dir / "existing-validation.json"
        report_path.write_text("do not replace\n", encoding="utf-8")
        result = self.run_cli(
            VALIDATOR,
            "--workbook",
            self.output_xlsx,
            "--source-json",
            self.input_json,
            "--output-json",
            report_path,
            "--mark-valid",
        )
        self.assertEqual(result.returncode, 2)
        workbook = load_workbook(self.output_xlsx)
        values = {
            row[0].value: row[1].value
            for row in workbook["Methodology"].iter_rows(min_row=2, max_col=2)
            if row[0].value
        }
        self.assertEqual(values["Validation Result"], "Pending")
        self.assertEqual(report_path.read_text(encoding="utf-8"), "do not replace\n")


def main() -> int:
    if OPENPYXL_IMPORT_ERROR is not None:
        print(
            "Missing dependency 'openpyxl'. Install it in the active Python environment "
            "(for example: python3 -m pip install openpyxl) and rerun.",
            file=sys.stderr,
        )
        return 2
    suite = unittest.defaultTestLoader.loadTestsFromTestCase(ContentPlanScriptsTest)
    result = unittest.TextTestRunner(verbosity=2).run(suite)
    return 0 if result.wasSuccessful() else 1


if __name__ == "__main__":
    raise SystemExit(main())
